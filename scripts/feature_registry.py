#!/usr/bin/env python3
"""Feature Registry: generazione, validazione e viste derivate.

Lo stato di una feature vive in `docs/roadmap/feature-registry.yaml` e in nessun altro posto.
Roadmap, Wiki e workbook referenziano il `feature_id`; questo script produce le viste derivate e
verifica che i riferimenti non siano rotti.

Nasce dallo stesso difetto che il repository ha gia' pagato quattro volte col conteggio dei test:
uno stato scritto a mano in due posti diverge, e la seconda copia diventa una bugia con la data
sbagliata. Qui la copia e' generata, e il generatore fallisce se la sorgente non regge.

Uso:
    python scripts/feature_registry.py validate           # gate: esce 1 se ci sono errori
    python scripts/feature_registry.py generate           # feature-registry.json + project-graph.json
    python scripts/feature_registry.py wiki               # blocchi di stato nel repo (docs/characters/)
    python scripts/feature_registry.py shortlist          # le cinque viste corte di docs/roadmap/
    python scripts/feature_registry.py report             # tabella di audit (markdown su stdout)
    python scripts/feature_registry.py suite --run-log Saved/Logs/RefactorTactics.log
                                                         # gate: ESEGUITI vs DICHIARATI, esce 1 se mancano (#486)
    python scripts/feature_registry.py deploy --wiki-root PATH --write   # blocchi + Stato-delle-feature nel clone
                                                                 # (senza --write e' sola lettura)

Opzioni comuni:
    --wiki-root PATH    radice del clone della Wiki. Dal 2026-08-10 (D-076) il clone e' la **fonte**
                        delle pagine di gioco: i ref `wiki:<PageName>` si risolvono solo li'.
    --check             `generate`/`wiki`/`shortlist`/`deploy` non scrivono: falliscono se disallineati
"""
import argparse
import json
import os
import re
import sys

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REGISTRY_YAML = os.path.join(REPO, "docs", "roadmap", "feature-registry.yaml")
REGISTRY_JSON = os.path.join(REPO, "docs", "roadmap", "feature-registry.json")
# Il contratto delle viste **non-feature**: diagnostica, epic/milestone, sedute, PIE, scenari.
# Sta a parte da `feature-registry.json` perche' quello e' l'owner delle feature e ha gia' altri
# consumatori: allargarlo a tutto ripeterebbe su di esso l'errore evitato sul `.yaml`.
PROJECT_GRAPH_JSON = os.path.join(REPO, "docs", "roadmap", "project-graph.json")
ROADMAP_V01 = os.path.join(REPO, "docs", "roadmap", "roadmap-v0.1.md")
ROADMAP_CHECKPOINT = os.path.join(REPO, "docs", "roadmap", "roadmap-checkpoint.md")
ROADMAP_POST_V01 = os.path.join(REPO, "docs", "roadmap", "roadmap-post-v0.1.md")
EXECUTION_GRAPH = os.path.join(REPO, "docs", "roadmap", "execution-graph.yaml")
EXECUTION_MAP_PAGE = os.path.join(REPO, "docs", "roadmap", "execution-map.md")
TESTS_DIR = os.path.join(REPO, "Source", "RefactorTactics", "Tests")
SCENARIOS_DIR = os.path.join(REPO, "Scenarios")

FEATURE_ID_RE = re.compile(r"^RT-FEAT-[A-Z0-9]+(-[A-Z0-9]+)*$")
GATE_NAMES = [
    "spec", "data", "runtime", "log_debug", "automation",
    "scenario", "ui_wiki", "packaged", "network_privacy",
    "replay_representable",
]
GATE_VALUES = {"done", "partial", "todo", "na"}

# Le release, **in ordine**, e l'insieme che le valida: una tupla sola invece di quattro elenchi
# paralleli. Fino al 2026-08-13 i valori ammessi erano `{v0.1, v0.2, future}` mentre
# `roadmap-post-v0.1.md` era gia' owner dichiarato di v0.2, v0.3 e v0.4: il registry non sapeva
# esprimere la roadmap che il repository aveva gia' scritto, e cinque feature stavano in `future`
# per assenza di un valore, non per assenza di una decisione. L'ordine viveva ricopiato in tre
# punti (`render_status_page` due volte, `render_shortlist_features` una) e aggiungere una release
# voleva dire trovarli tutti e tre: chi ne mancava uno otteneva una vista che ometteva in silenzio
# una release intera, che e' esattamente il difetto che §3 di `roadmap.shortlist.md` aveva.
RELEASE_ORDER = ("v0.1", "v0.2", "v0.3", "v0.4", "future")
RELEASES = set(RELEASE_ORDER)
PRIORITIES = {"P0", "P1", "P2", "P3"}
KINDS = {"gameplay", "ui", "tooling", "data", "infra", "content"}

# I `kind` che la Wiki documenta davvero. La Wiki e' rivolta a chi gioca — meccaniche, personaggi,
# fazioni — e non parla del map editor ne' dell'asset pipeline: e' una scelta gia' presa e gia'
# visibile nei dati, non una regola nuova. Al 2026-08-10 le 60 feature con una pagina sono tutte
# `gameplay`, `ui` o `content`, e nessuna delle 18 `tooling`/`data`/`infra` ne ha una.
# Pretenderla da loro produceva 18 avvisi su 27 che non si sarebbero chiusi mai — e la regola 5 di
# `feature-registry.md` gia' dichiara legittimo quel caso: «una feature interna puo' avere la UI e
# nessuna pagina (`partial` anche li')». Un avviso che contraddice il proprio owner documentale non
# denuncia una lacuna: insegna a ignorare anche gli altri avvisi.
WIKI_DOCUMENTED_KINDS = {"gameplay", "ui", "content"}

# Ordine di maturita'. `DEFERRED` e `BLOCKED` sono fuori scala: dichiarano una decisione,
# non un grado di completezza, e per questo il controllo di coerenza li salta.
STATUS_ORDER = [
    "IDEA", "DESIGNED", "SPECIFIED", "IMPLEMENTING",
    "TESTABLE", "INTEGRATED", "RELEASE_READY", "DONE",
]
STATUS_OFF_SCALE = {"DEFERRED", "BLOCKED"}
ALL_STATUS = set(STATUS_ORDER) | STATUS_OFF_SCALE

# Un gate `na` conta come soddisfatto: «non applicabile» e' una risposta, non un buco.
# `partial` non lo e' mai: e' il modo in cui una feature dichiara di non aver finito.
SATISFIED = {"done", "na"}

MARKER_BEGIN = "<!-- RT_FEATURE_STATUS:BEGIN {fid} -->"
MARKER_END = "<!-- RT_FEATURE_STATUS:END {fid} -->"
MARKER_ANY = re.compile(
    r"<!-- RT_FEATURE_STATUS:BEGIN (?P<fid>[A-Za-z0-9\-]+) -->.*?"
    r"<!-- RT_FEATURE_STATUS:END (?P=fid) -->",
    re.S,
)


# ---------------------------------------------------------------------------
# Lettura della sorgente e delle evidenze
# ---------------------------------------------------------------------------

def load_registry():
    try:
        import yaml
    except ImportError:
        sys.exit("Serve PyYAML: pip install pyyaml")
    with open(REGISTRY_YAML, encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def known_tests():
    """I nomi dei test sono la prova di cio' che esiste: si misurano, non si citano."""
    names = set()
    if not os.path.isdir(TESTS_DIR):
        return names
    pattern = re.compile(r'"(RefactorTactics\.[A-Za-z0-9_.]+)"')
    for entry in sorted(os.listdir(TESTS_DIR)):
        if not entry.endswith(".cpp"):
            continue
        with open(os.path.join(TESTS_DIR, entry), encoding="utf-8", errors="replace") as fh:
            names.update(pattern.findall(fh.read()))
    return names


def known_scenarios():
    """ScenarioId dichiarati nei file, non dedotti dal percorso (scenario-index-e-tag.md §3)."""
    ids = set()
    for root, _dirs, files in os.walk(SCENARIOS_DIR):
        for name in files:
            if not name.endswith(".json") or name.startswith("_"):
                continue
            path = os.path.join(root, name)
            try:
                with open(path, encoding="utf-8-sig") as fh:
                    data = json.load(fh)
            except (json.JSONDecodeError, OSError):
                continue
            sid = data.get("scenarioId")
            if sid:
                ids.add(sid)
    return ids


def known_roadmap_refs():
    """Epic e checkpoint dichiarati dalla roadmap di release; milestone da quella di esecuzione.

    Riconosce sia la forma prefissata `E9.1` (convenzione dal 2026-08-08) sia `CP 9.1` senza
    prefisso, che resta valida nel corpus storico e si legge come checkpoint di **epic**.
    """
    epics, checkpoints, milestones = set(), set(), set()
    if os.path.isfile(ROADMAP_V01):
        text = open(ROADMAP_V01, encoding="utf-8").read()
        epics.update(re.findall(r"^### (E\d+) —", text, re.M))
        epics.update(re.findall(r"\*\*(E\d+)\*\*", text))
        checkpoints.update(re.findall(r"CP (\d+\.\d+)", text))
        checkpoints.update(re.findall(r"^\| \*\*(\d+\.\d+)\*\*", text, re.M))
        # Forma prefissata: `E9.1` dichiara il checkpoint 9.1 dell'epic E9.
        checkpoints.update(re.findall(r"\bE(\d+\.\d+)\b", text))
    if os.path.isfile(ROADMAP_CHECKPOINT):
        text = open(ROADMAP_CHECKPOINT, encoding="utf-8").read()
        milestones.update(re.findall(r"\*\*(M\d+)\*\*", text))
    return epics, checkpoints, milestones


def test_pattern_matches(pattern, names):
    """`Actions.Wait` vale per tutto il sottoalbero `Actions.Wait.*`.

    I nomi dei test hanno profondita' variabile e la regola del progetto vieta che un nome sia
    prefisso di un altro (§6.1 della roadmap): un riferimento a un ramo non e' quindi ambiguo,
    e obbligare a scrivere `.*` ovunque renderebbe il registry rumoroso senza guadagno.
    """
    if pattern.endswith("*"):
        prefix = pattern[:-1]
        return any(n.startswith(prefix) for n in names)
    return pattern in names or any(n.startswith(pattern + ".") for n in names)


def scenario_lists(feature):
    """`scenarios` accetta una lista piatta oppure {planned: [...]} per cio' che non esiste ancora."""
    raw = feature.get("scenarios") or []
    if isinstance(raw, dict):
        return list(raw.get("present", []) or []), list(raw.get("planned", []) or [])
    present, planned = [], []
    for item in raw:
        if isinstance(item, dict) and "planned" in item:
            planned.extend(item["planned"] or [])
        else:
            present.append(item)
    return present, planned


# ---------------------------------------------------------------------------
# Derivazione dello stato dai gate
# ---------------------------------------------------------------------------

def derive_status(gates):
    """Lo stato non e' un'opinione: e' una funzione dei gate.

    Restituisce il gradino piu' alto che i gate reggono. Il validator confronta questo valore
    con quello dichiarato: se il dichiarato e' piu' alto, e' un errore — e' esattamente il modo
    in cui una feature finisce per essere «Done» perche' si vede in PIE.
    """
    def ok(*names):
        return all(gates.get(n) in SATISFIED for n in names)

    core = ("spec", "data", "runtime", "log_debug", "automation")
    # `replay_representable` sta nel gradino DONE accanto a `packaged` e `network_privacy`, e non
    # piu' in basso, per una ragione misurata: al momento in cui e' entrato (2026-08-12) una sola
    # feature era DONE, quindi il gate non poteva far regredire nessuno stato dichiarato — e un
    # gate nuovo che retrocede meta' registry al primo giro non viene creduto, viene aggirato.
    # E' anche il posto giusto nel merito: e' una proprieta' trasversale che si verifica per ultima,
    # esattamente come le altre due di questo gradino.
    if ok(*core, "scenario", "ui_wiki", "packaged", "network_privacy", "replay_representable"):
        return "DONE"
    if ok(*core, "scenario", "ui_wiki"):
        return "RELEASE_READY"
    if ok(*core, "scenario"):
        return "INTEGRATED"
    if ok("spec", "runtime", "automation"):
        return "TESTABLE"
    if gates.get("runtime") in ("done", "partial"):
        return "IMPLEMENTING"
    if gates.get("spec") == "done":
        return "SPECIFIED"
    if gates.get("spec") == "partial":
        return "DESIGNED"
    return "IDEA"


def gate_progress(gates):
    """«6/8 gate» invece di «73%»: il primo si verifica, il secondo si contratta."""
    applicable = [g for g in GATE_NAMES if gates.get(g) != "na"]
    done = [g for g in applicable if gates.get(g) == "done"]
    return len(done), len(applicable)


# ---------------------------------------------------------------------------
# Validazione
# ---------------------------------------------------------------------------

def github_config(registry):
    """`owner`/`repository`/`branch` da cui si derivano tutti i link verso GitHub."""
    return ((registry.get("meta") or {}).get("project") or {}).get("github") or {}


def git_remote_slug():
    """`owner/repo` del remote `origin`, se c'e'. Legge la config locale, non la rete."""
    try:
        import subprocess
        out = subprocess.run(["git", "remote", "get-url", "origin"], cwd=REPO,
                             capture_output=True, text=True, timeout=10)
        match = re.search(r"[:/]([^/:]+)/([^/]+?)(?:\.git)?\s*$", out.stdout.strip())
        return f"{match.group(1)}/{match.group(2)}" if match else None
    except Exception:
        return None


def validate(registry, wiki_root=None, editor_ctx=None):
    errors, warnings = [], []
    features = registry.get("features") or []

    # Misurato una volta e riusato: qui per `pie_refs`, e piu' sotto da `validate_editor_sessions`.
    # Chi ha gia' il contesto (`build_graph`) lo passa e non si rilegge nulla.
    editor_ctx = editor_ctx or editor_context()
    pie_registry = editor_ctx["pie"]

    github = github_config(registry)
    missing = [k for k in ("owner", "repository", "branch") if not github.get(k)]
    if missing:
        errors.append("meta.project.github incompleto: mancano " + ", ".join(missing)
                      + " — senza questi nessun link a GitHub e' derivabile")
    else:
        slug = git_remote_slug()
        declared = f"{github['owner']}/{github['repository']}"
        if slug and slug.lower() != declared.lower():
            # Avviso e non errore: su un fork la divergenza e' legittima, e un gate che fallisce
            # su ogni fork verrebbe disattivato — cioe' non varrebbe piu' nemmeno qui.
            warnings.append(f"meta.project.github dichiara {declared} ma il remote e' {slug}")
    tests = known_tests()
    scenarios = known_scenarios()
    epics, checkpoints, milestones = known_roadmap_refs()
    ids = [f.get("feature_id") for f in features]

    for fid in sorted({i for i in ids if ids.count(i) > 1}):
        errors.append(f"FeatureId duplicato: {fid}")

    id_set = set(ids)
    testable_or_more = set(STATUS_ORDER[STATUS_ORDER.index("TESTABLE"):])
    claimed_scenarios = set()

    for feature in features:
        fid = feature.get("feature_id", "<senza id>")
        where = f"[{fid}]"

        if not FEATURE_ID_RE.match(fid or ""):
            errors.append(f"{where} feature_id non conforme a RT-FEAT-<AREA>-<NOME>")

        status = feature.get("status")
        if status not in ALL_STATUS:
            errors.append(f"{where} status non valido: {status!r}")
        if feature.get("release") not in RELEASES:
            errors.append(f"{where} release non valida: {feature.get('release')!r}")
        if feature.get("priority") not in PRIORITIES:
            errors.append(f"{where} priority non valida: {feature.get('priority')!r}")
        if feature.get("kind") not in KINDS:
            errors.append(f"{where} kind non valido: {feature.get('kind')!r}")

        gates = feature.get("gates") or {}
        for gate in GATE_NAMES:
            value = gates.get(gate)
            if value is None:
                errors.append(f"{where} gate mancante: {gate}")
            elif value not in GATE_VALUES:
                errors.append(f"{where} gate {gate} non valido: {value!r}")
        for gate in gates:
            if gate not in GATE_NAMES:
                errors.append(f"{where} gate sconosciuto: {gate}")

        if status in STATUS_ORDER and all(gates.get(g) in GATE_VALUES for g in GATE_NAMES):
            derived = derive_status(gates)
            if status != derived:
                missing = [g for g in GATE_NAMES if gates.get(g) not in SATISFIED]
                if STATUS_ORDER.index(status) > STATUS_ORDER.index(derived):
                    detail = f"gate non completati: {', '.join(missing)}"
                else:
                    detail = "i gate dicono che e' piu' avanti di quanto dichiari"
                errors.append(
                    f"{where} status {status} diverge dai gate (derivato: {derived}); {detail}"
                )

        last = feature.get("last_verified") or {}
        if status in testable_or_more and not (last.get("date") and last.get("commit")):
            errors.append(f"{where} last_verified assente ma status e' {status}")

        roadmap = feature.get("roadmap") or {}
        epic = roadmap.get("epic")
        if epic and epic not in epics:
            errors.append(f"{where} epic inesistente nella roadmap di release: {epic}")
        milestone = roadmap.get("milestone")
        if milestone and milestone not in milestones:
            errors.append(f"{where} milestone inesistente nella roadmap di esecuzione: {milestone}")
        for key in roadmap or {}:
            if key not in ("epic", "milestone", "checkpoints", "out_of_release_scope"):
                errors.append(f"{where} campo roadmap sconosciuto: {key}")
        for cp in roadmap.get("checkpoints") or []:
            if str(cp) not in checkpoints:
                container = epic or milestone
                label = f"{container}.{str(cp).split('.', 1)[-1]}" if container else f"CP {cp}"
                errors.append(f"{where} checkpoint inesistente: {label}")
        if roadmap.get("checkpoints") and not (epic or milestone):
            errors.append(f"{where} checkpoint dichiarati senza epic ne' milestone: "
                          "il riferimento non sarebbe risolvibile")
        if roadmap.get("out_of_release_scope") and epic:
            errors.append(f"{where} ha un'epic e insieme out_of_release_scope: decidi quale dei due")

        for dep in feature.get("dependencies") or []:
            if dep not in id_set:
                errors.append(f"{where} dipendenza verso FeatureId inesistente: {dep}")

        # `completed_by` chiude il doppio conteggio: quando un pezzo di questa feature e' portato
        # da un'altra, il rimando e' un DATO e non una frase nelle note. Senza, la stessa mancanza
        # risulta due volte e chiudendola bisogna ricordarsi di aggiornare due righe — che e'
        # precisamente il meccanismo che questo registry esiste per eliminare.
        for other in feature.get("completed_by") or []:
            if other not in id_set:
                errors.append(f"{where} completed_by verso FeatureId inesistente: {other}")
            elif other == fid:
                errors.append(f"{where} completed_by verso se stessa")
        if feature.get("completed_by") and not [
                g for g in GATE_NAMES if gates.get(g) in ("partial", "todo")]:
            warnings.append(
                f"{where} dichiara completed_by ma non ha gate aperti: "
                "se non manca niente, il rimando e' vecchio")

        for spec in feature.get("owner_specs") or []:
            if not os.path.isfile(os.path.join(REPO, spec)):
                errors.append(f"{where} owner spec inesistente: {spec}")

        for pattern in feature.get("tests") or []:
            if not test_pattern_matches(pattern, tests):
                errors.append(f"{where} test ref senza corrispondenza nella suite: {pattern}")

        present, planned = scenario_lists(feature)
        claimed_scenarios.update(present)
        for sid in present:
            if sid not in scenarios:
                errors.append(f"{where} ScenarioId inesistente: {sid}")
        for sid in planned:
            if sid in scenarios:
                # Errore e non avviso: uno scenario marcato `planned` che esiste davvero non e' un
                # promemoria, e' il registry che dice una cosa falsa. Un avviso qui si sarebbe
                # accumulato fino a diventare rumore — che e' il modo in cui il conteggio dei test
                # e' divergito cinque volte.
                errors.append(
                    f"{where} scenario dichiarato planned ma presente in Scenarios/: {sid} — "
                    "promuovilo fra gli scenari presenti e rivedi il gate `scenario`"
                )

        for ref in feature.get("wiki_refs") or []:
            if ref.startswith("wiki:"):
                # Senza `wiki_root` il clone non e' leggibile e il riferimento non si controlla.
                # Non e' un errore — non tutti hanno il clone accanto — ma nemmeno un silenzio: il
                # comando `validate` conta questi ref e lo dice, perche' un verde su qualcosa che
                # non e' stato aperto e' peggio di un rosso.
                if wiki_root and not wiki_page_by_name(wiki_root, ref[len("wiki:"):]):
                    errors.append(f"{where} pagina Wiki inesistente nel clone: {ref}")
            elif not os.path.isfile(os.path.join(REPO, ref)):
                errors.append(f"{where} pagina Wiki sorgente inesistente: {ref}")

        # --- il gate `packaged` e la verifica che lo dimostra ---
        #
        # `packaged` dice «verificata nella build packaged della release corrente», e chi la verifica
        # e' una voce `PIE-*` eseguita da una persona. Fino al 2026-08-10 quel legame non era un
        # dato: viveva nel corpo delle issue GitHub («Verifica PIE `PIE-V01-HUD`»), fuori dal
        # repository e fuori da ogni controllo. La conseguenza non e' teorica — il giorno in cui
        # `PIE-V01-HUD` diventa verde, nessuno sa che quattro feature possono avanzare il gate.
        pie_refs = feature.get("pie_refs") or []
        for ref in pie_refs:
            if ref not in pie_registry:
                errors.append(f"{where} voce PIE inesistente nel registro: {ref}")
        # Tutte le voci, comprese quelle che il registro non conosce: una voce inesistente vale
        # `None`, che non e' verde. Filtrarle qui — `if r in pie_registry` — le avrebbe rese
        # invisibili a questi due controlli, e una lista `[voce_verde, voce_inesistente]` avrebbe
        # dato «tutte verdi»: il gate sarebbe potuto restare `done` senza che nessuno lo negasse,
        # e l'avviso «puo' avanzare» sarebbe comparso su una prova che non esiste.
        cited = [pie_registry.get(r) for r in pie_refs]
        if gates.get("packaged") == "done" and pie_refs and not all(s == "✅" for s in cited):
            aperte = [r for r in pie_refs if pie_registry.get(r) not in ("✅",)]
            errors.append(
                f"{where} gate packaged=done ma le voci che lo dimostrano non sono verdi: "
                + ", ".join(f"{r} {pie_registry.get(r) or '❓'}" for r in aperte)
                + " — il gate afferma una verifica che il registro PIE non conferma")
        if gates.get("packaged") in ("partial", "todo") and pie_refs and all(s == "✅" for s in cited):
            warnings.append(
                f"{where} tutte le voci PIE che la dimostrano sono verdi ({', '.join(pie_refs)}) "
                f"ma il gate packaged e' ancora {gates.get('packaged')}: puo' avanzare")

        # `ui_wiki` conta due meta' (regola 5 di `feature-registry.md`): «spiegata all'utente» e
        # «leggibile in gioco». `done` le dichiara ENTRAMBE, quindi senza una pagina collegata il
        # gate afferma qualcosa che nessun dato sostiene. `partial` invece e' legittimo senza
        # pagina — e' precisamente il caso «la UI c'e', la pagina no» che la regola prevede.
        if gates.get("ui_wiki") == "done" and not (feature.get("wiki_refs") or []):
            errors.append(
                f"{where} gate ui_wiki=done ma nessuna wiki_refs: `done` dichiara entrambe le meta' "
                "(spiegata all'utente e leggibile in gioco) e la prima non ha nulla che la dimostri. "
                "Collega la pagina, oppure il gate e' `partial`")

        # --- warning: lacune che non rompono nulla ma vanno viste ---
        if feature.get("kind") in WIKI_DOCUMENTED_KINDS and not (feature.get("wiki_refs") or []):
            warnings.append(f"{where} nessuna pagina Wiki collegata")
        if status in testable_or_more and feature.get("kind") == "gameplay" and not present:
            warnings.append(f"{where} feature di gameplay {status} senza scenario che la dimostri")
        if status == "SPECIFIED" and not (feature.get("issues") or []) and not epic and not milestone:
            warnings.append(f"{where} SPECIFIED senza issue ne' assegnazione in roadmap")
        if planned:
            warnings.append(
                f"{where} {len(planned)} scenario/i dichiarati planned: "
                + ", ".join(planned)
            )

    # --- il verso opposto: nessuno scenario senza feature che lo rivendichi ---
    #
    # Il controllo sopra va da feature a scenario ("lo ScenarioId dichiarato esiste?"). Mancava
    # quello inverso, e la conseguenza non e' teorica: al 2026-08-09 erano 6 su 54 gli scenari che
    # nessuna feature rivendicava - fra cui `Visual.Map.HighCoverBlocks` e
    # `Spec.Environment.ElectricPropagation`, entrambi documentati in `scenario-map.md` e nel corpus
    # visuale. Uno scenario che nessuno rivendica si esegue, passa, e non dimostra niente a nessuno:
    # e' la stessa famiglia del dato che nessun consumatore legge.
    #
    # Errore e non avviso, per simmetria con "scenario dichiarato planned ma presente": in entrambi
    # i casi il registry sta dicendo qualcosa di falso sulla copertura, non segnalando una lacuna.
    orphans = sorted(set(scenarios) - claimed_scenarios)
    for sid in orphans:
        errors.append(
            f"ScenarioId non rivendicato da nessuna feature: {sid} — "
            "aggiungilo agli `scenarios` della feature che dimostra, oppure spiega in `notes` "
            "perche' non appartiene a nessuna"
        )

    # NB: «quanti riferimenti `wiki:` non ho potuto verificare» **non** entra fra i warning, e per
    # questo non si calcola qui. I warning finiscono in `project-graph.json`, che e' versionato e ha
    # un `--check`: un avviso che dipende dalla presenza di `--wiki-root` renderebbe il file generato
    # diverso a seconda di come lo lanci, e il `--check` rosso a giorni alterni. Quell'avviso
    # descrive quanto ha guardato **l'esecuzione**, non lo stato del progetto: lo calcola e lo stampa
    # il comando `validate`, che sa con che flag e' partito.

    # --- riferimenti a feature dalle pagine gia' generate ---
    for page, fids in wiki_blocks_in_repo().items():
        for fid in fids:
            if fid not in id_set:
                errors.append(f"[{page}] blocco di stato per FeatureId inesistente: {fid}")

    # --- l'altra sorgente delle *map, con le stesse pretese di questa ---
    session_errors, session_warnings = validate_editor_sessions(editor_ctx)
    errors += session_errors
    warnings += session_warnings

    exec_errors, exec_warnings = validate_execution_graph(editor_ctx, registry)
    errors += exec_errors
    warnings += exec_warnings

    return errors, warnings


def wiki_blocks_in_repo():
    """Pagine del repo che contengono gia' un blocco generato, con i FeatureId citati."""
    found = {}
    # Solo `docs/characters/`: da D-076 le pagine di gioco stanno nel clone, e i loro blocchi li
    # verifica `deploy`.
    for base in ("docs/characters",):
        root_dir = os.path.join(REPO, base)
        for root, _dirs, files in os.walk(root_dir):
            for name in files:
                if not name.endswith(".md"):
                    continue
                path = os.path.join(root, name)
                text = open(path, encoding="utf-8", errors="replace").read()
                fids = MARKER_ANY.findall(text)
                fids = re.findall(r"RT_FEATURE_STATUS:BEGIN ([A-Za-z0-9\-]+)", text)
                if fids:
                    found[os.path.relpath(path, REPO).replace("\\", "/")] = fids
    return found


# ---------------------------------------------------------------------------
# Viste derivate
# ---------------------------------------------------------------------------

def jsonable(value):
    """PyYAML restituisce `datetime.date` per `2026-08-08`: nel JSON serve la stringa ISO."""
    import datetime
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.isoformat()
    if isinstance(value, dict):
        return {k: jsonable(v) for k, v in value.items()}
    if isinstance(value, list):
        return [jsonable(v) for v in value]
    return value


def build_json(registry):
    features = []
    for feature in registry.get("features") or []:
        gates = feature.get("gates") or {}
        done, total = gate_progress(gates)
        present, planned = scenario_lists(feature)
        roadmap = feature.get("roadmap") or {}
        features.append({
            "feature_id": feature.get("feature_id"),
            "title": feature.get("title"),
            "area": feature.get("area"),
            "kind": feature.get("kind"),
            "release": feature.get("release"),
            "priority": feature.get("priority"),
            "status": feature.get("status"),
            "status_derived": derive_status(gates) if all(
                gates.get(g) in GATE_VALUES for g in GATE_NAMES) else None,
            "gates": {g: gates.get(g) for g in GATE_NAMES},
            "gates_done": done,
            "gates_applicable": total,
            "roadmap": {
                "epic": roadmap.get("epic"),
                "milestone": roadmap.get("milestone"),
                "checkpoints": [str(c) for c in (roadmap.get("checkpoints") or [])],
                "out_of_release_scope": (roadmap.get("out_of_release_scope") or "").strip(),
            },
            "dependencies": feature.get("dependencies") or [],
            "completed_by": feature.get("completed_by") or [],
            "owner_specs": feature.get("owner_specs") or [],
            "issues": feature.get("issues") or [],
            "tests": feature.get("tests") or [],
            "scenarios": present,
            "scenarios_planned": planned,
            "wiki_refs": feature.get("wiki_refs") or [],
            "wiki_note": (feature.get("wiki_note") or "").strip(),
            "pie_refs": feature.get("pie_refs") or [],
            "last_verified": jsonable(feature.get("last_verified") or {}),
            "notes": (feature.get("notes") or "").strip(),
        })
    return {
        "_generated": "NON EDITARE: generato da docs/roadmap/feature-registry.yaml "
                      "con scripts/feature_registry.py generate",
        "meta": jsonable(registry.get("meta") or {}),
        "project": jsonable((registry.get("meta") or {}).get("project") or {}),
        "count": len(features),
        "features": features,
    }


def roadmap_ref(roadmap):
    """Riferimento di roadmap nella forma prefissata.

    Un checkpoint non si cita mai senza il proprio contenitore: `CP 10.1` e' «Activate e Interact»
    in E10 **e** «listen server» in M10, e `CP 6.1` e' due cose diverse in E2 e in E6. Il prefisso
    non e' decorazione, e' cio' che rende il riferimento risolvibile.
    """
    container = roadmap.get("epic") or roadmap.get("milestone")
    if not container:
        return "fuori scope" if roadmap.get("out_of_release_scope") else "—"
    checkpoints = roadmap.get("checkpoints") or []
    if not checkpoints:
        return container
    # Il checkpoint prefissato contiene gia' il contenitore: ripeterlo darebbe «E14 · E14.4».
    return ", ".join(f"{container}.{c.split('.', 1)[-1]}" for c in checkpoints)


def status_block(entry):
    roadmap = entry["roadmap"]
    ref = roadmap_ref(roadmap)
    scenario = entry["scenarios"][0] if entry["scenarios"] else (
        entry["scenarios_planned"][0] + " (pianificato)" if entry["scenarios_planned"] else "—")
    verified = entry["last_verified"] or {}
    # La marcatura e' proporzionale allo scarto fra cio' che la pagina descrive e cio' che il gioco
    # fa. Una forma uniforme per una feature RELEASE_READY e per una con 1 gate su 9 viene letta
    # come boilerplate e saltata — e questa Wiki e' pubblica: chi arriva da una ricerca legge le
    # regole, non il riquadro.
    status = entry["status"]
    if status in ("IDEA", "DESIGNED", "SPECIFIED"):
        intro = ("> ⚠️ **Progettata, non implementata.** Questa pagina descrive una meccanica "
                 "**decisa e documentata** che il gioco **non esegue ancora**: oggi non è "
                 "giocabile. Blocco generato dal Feature Registry, non modificare a mano.  ")
    elif status in ("IMPLEMENTING", "TESTABLE"):
        intro = ("> 🚧 **Parzialmente giocabile.** Il codice esiste ma la feature non è completa: "
                 "i gate qui sotto dicono quanto manca. Blocco generato dal Feature Registry, "
                 "non modificare a mano.  ")
    else:
        intro = "> **Stato di sviluppo** — generato dal Feature Registry, non modificare a mano.  "

    lines = [
        MARKER_BEGIN.format(fid=entry["feature_id"]),
        "",
        intro,
        f"> Feature: `{entry['feature_id']}` · Release: `{entry['release']}` · Roadmap: `{ref}`  ",
        f"> Stato: **{entry['status']}** · Gate: `{entry['gates_done']}/{entry['gates_applicable']}`  ",
        f"> Scenario: `{scenario}`  ",
    ]
    # La nota e' l'unico testo libero del blocco: serve a non perdere il dettaglio che i banner
    # scritti a mano portavano («esiste il dato ma nessun eroe lo usa»), tenendolo pero' in un
    # posto solo. Resta una riga: se serve un paragrafo, va nel corpo della pagina.
    if entry.get("completed_by"):
        lines.append("> I pezzi che mancano li porta: "
                     + " · ".join(f"`{c}`" for c in entry["completed_by"]) + "  ")
    if entry.get("wiki_note"):
        lines.append(f"> {entry['wiki_note']}  ")
    lines += [
        f"> Verificato il `{verified.get('date', '—')}` su `{verified.get('commit', '—')}`",
        "",
        MARKER_END.format(fid=entry["feature_id"]),
    ]
    return "\n".join(lines)


def render_execution_mermaid(execution, nodes, label):
    """L'execution graph come diagramma Mermaid, dentro la pagina Wiki che GitHub rende.

    Il Control Center disegna un SVG proprio, ma vive in locale. Qui serve una figura che stia in
    una pagina markdown senza binari da committare: la stessa scelta che `roadmap-v0.1.md` ha gia'
    fatto per il grafo delle epic.

    ⚠️ **Non fidarti di un headless anonimo per dire se rende.** Chrome headless senza sessione
    mostra «Unable to render rich display» su qualunque diagramma — anche su `A[uno] --> B[due]`,
    66 byte — perche' non esegue il rendering client-side di GitHub. Su quella misura ho creduto
    rotta una figura che si e' sempre vista, ho semplificato tre volte le etichette e ho spostato
    l'intera pagina nel repository per rimediare a un difetto che non esisteva. La verifica valida
    e' un browser vero.

    Tre cose che il diagramma deve dire, e come le dice:

    - **le corsie**: la lane si legge dalla FORMA del nodo. Un `subgraph` per lane sembrava la
      scelta ovvia e non lo era: Mermaid dispone i blocchi dove vuole, e nella prima stesura `PIE`
      finiva a sinistra, `CODE` al centro e `ASSET` a destra, con gli archi di evidenza a
      traversare tutta la figura;
    - **il tipo di legame**: freccia piena per una dipendenza dura, tratteggiata per l'ordine
      consigliato e la navigazione. Un `follows` che sembrasse un blocco sarebbe la bugia che il
      modello esiste per evitare;
    - **i lati liberi**: un cerchietto attaccato al nodo — dove si comincia (nessun ingresso) e
      dove una catena si esaurisce (nessuna uscita).

    Gli id Mermaid non ammettono `#` o `:`: si normalizzano, e l'etichetta resta leggibile.
    """
    ident = lambda nid: re.sub(r"[^A-Za-z0-9]", "_", str(nid))
    edges = execution.get("edges") or []
    visible = {n["id"] for n in execution.get("nodes") or []}
    hard_in, hard_out = {}, {}
    for edge in edges:
        if edge.get("hard") and edge["from"] in visible and edge["to"] in visible:
            hard_in.setdefault(edge["to"], []).append(edge["from"])
            hard_out.setdefault(edge["from"], []).append(edge["to"])

    # La lane si legge dalla FORMA, non da un `subgraph`. Mermaid dispone i subgraph come blocchi
    # separati e non come corsie: nella prima stesura `PIE` finiva a sinistra, `CODE` al centro e
    # `ASSET` a destra, con gli archi di evidenza a traversare tutta la figura. La forma viaggia
    # col nodo e non impone un layout.
    shape = {"code": ("[", "]"), "asset": ("[[", "]]"), "pie": ("([", "])")}
    lines = ["", "```mermaid", "flowchart LR"]
    for node in execution.get("nodes") or []:
        nid = ident(node["id"])
        open_b, close_b = shape.get(node.get("execution_lane"), ("[", "]"))
        # Il glifo di stato precede l'etichetta solo quando esiste, o il nodo nasce con uno spazio
        # davanti al `#` — visibile nella figura resa.
        #
        # ⚠️ Fra il 2026-08-13 e la correzione, questa riga e' stata impoverita a `164 READY` —
        # niente glifo, niente `#`, niente `<br/>` — inseguendo un difetto che **non esisteva**:
        # Chrome headless anonimo mostra «Unable to render rich display» su qualunque Mermaid,
        # anche su `A[uno] --> B[due]`, perche' non esegue il rendering client-side di GitHub. Su
        # un browser vero il diagramma si e' sempre visto. Le tre semplificazioni non hanno mai
        # corretto niente, e la lezione sta nel commit: prima di semplificare, prova lo strumento
        # di misura su un caso che sai buono.
        head = " ".join(x for x in (node.get("state"), label(node["id"])) if x)
        lines.append(f'  {nid}{open_b}"{head}<br/>{node.get("readiness")}"{close_b}')
        # Il moncone: un cerchio con dentro un punto. Un nodo `(( ))` vuoto rende un pallino di
        # pochi pixel che alla scala della pagina sparisce, e un lato libero invisibile e' un lato
        # libero non detto.
        if not hard_in.get(node["id"]):
            lines.append(f"  i_{nid}((.)) -.- {nid}")
        if not hard_out.get(node["id"]):
            lines.append(f"  {nid} -.- o_{nid}((.))")

    for edge in edges:
        if edge["from"] not in visible or edge["to"] not in visible:
            continue
        a, b = ident(edge["from"]), ident(edge["to"])
        if edge.get("hard"):
            lines.append(f"  {a} ==> {b}")
        elif edge["type"] in ("follows", "related"):
            lines.append(f'  {a} -. {edge["type"]} .-> {b}')
        elif edge["type"] in ("implements", "verifies"):
            lines.append(f'  {a} -. {edge["type"]} .-> {b}')

    # Il colore non e' l'unico portatore: la readiness e' scritta dentro ogni nodo. Serve a
    # scorrere la figura, non a decodificarla.
    lines += [
        "  classDef ready fill:#173a2a,stroke:#4ec98a,color:#e6e9ef;",
        "  classDef blocked fill:#3a1f22,stroke:#e06c75,color:#e6e9ef;",
        "  classDef waiting fill:#3a3320,stroke:#e0b050,color:#e6e9ef;",
        "  classDef done fill:#1e2b24,stroke:#4ec98a,color:#99a1b3;",
        "  classDef unknown fill:#22262f,stroke:#4a5162,color:#99a1b3;",
        "  classDef stub fill:#0f1115,stroke:#5a6273;",
    ]
    by_class = {}
    for node in execution.get("nodes") or []:
        key = {"READY": "ready", "BLOCKED": "blocked", "DONE": "done",
               "WAITING_FOR_PIE": "waiting", "WAITING_FOR_ASSET": "waiting"}.get(
                   node.get("readiness"), "unknown")
        by_class.setdefault(key, []).append(ident(node["id"]))
    stubs = []
    for node in execution.get("nodes") or []:
        nid = ident(node["id"])
        if not hard_in.get(node["id"]):
            stubs.append(f"i_{nid}")
        if not hard_out.get(node["id"]):
            stubs.append(f"o_{nid}")
    if stubs:
        by_class["stub"] = stubs
    for key, ids in by_class.items():
        lines.append(f"  class {','.join(ids)} {key};")
    lines += ["```", ""]
    lines += [
        "**Come si legge.** Freccia piena `==>` una dipendenza dura: finche' l'origine non e' fatta,",
        "la destinazione non puo' cominciare. Freccia tratteggiata un legame che **non blocca** —",
        "`follows` e' ordine consigliato, `related` navigazione, `implements`/`verifies` dicono chi",
        "realizza e chi giudica. Il cerchietto attaccato a un nodo e' un **lato libero**: a sinistra",
        "vuol dire che niente lo trattiene, a destra che non sblocca niente di dichiarato.",
        "",
        "La **forma** dice chi deve intervenire: rettangolo `CODE` (repository), rettangolo doppio",
        "`ASSET` (un asset da costruire e committare), rettangolo arrotondato `PIE` (un verdetto",
        "davanti all'editor). Il colore ripete la readiness, che e' scritta dentro ogni nodo: serve",
        "a scorrere la figura, non a decodificarla.",
    ]
    return lines


def render_execution_map_page(graph):
    """`docs/roadmap/execution-map.md`: la figura delle dipendenze, dove GitHub la rende.

    Sta nel repository e non nella Wiki per un fatto misurato, non per gusto: le pagine Wiki non
    renderizzano Mermaid — nemmeno un diagramma di tre nodi — mentre i file versionati si'.
    """
    execution = (graph or {}).get("execution")
    if not execution:
        return None
    nodes = {n["id"]: n for n in execution.get("nodes") or []}

    def label(nid):
        if str(nid).startswith("capability:"):
            return str(nid).split(":", 1)[1]
        node = nodes.get(nid) or {}
        return f"#{node['ref']}" if node.get("kind") == "issue" else node.get("ref") or nid

    stats = execution.get("stats") or {}
    out = [
        "# Execution map — cosa blocca cosa",
        "",
        "> `GENERATO` · lo riscrive `python scripts/feature_registry.py shortlist`.",
        "> **Cosa e'**: la figura delle dipendenze di esecuzione, con lo stato di ogni nodo.",
        "> **Cosa non e'**: una fonte. Nodi, archi e readiness vengono da",
        "> [`project-graph.json`](project-graph.json), che li deriva dagli owner; la topologia da",
        "> [`execution-graph.yaml`](execution-graph.yaml).",
        "",
        "La stessa figura, interattiva e filtrabile, e' il tab **Execution Map** del",
        "[Project Control Center](../control-center/README.md). I conteggi e le tabelle stanno",
        "sulla Wiki, in [Stato del progetto]"
        "(https://github.com/DegrassiAaron/refactor-tactics-main/wiki/Stato-del-progetto):",
        "li' il diagramma non compare perche' **le pagine Wiki non rendono Mermaid**.",
        "",
        f"Fetta «{execution.get('status')}» · **{stats.get('nodes')} nodi** · "
        f"**{stats.get('hard_edges')} dipendenze dure** · "
        f"**{stats.get('soft_edges')} relazioni molli**.",
    ]
    out += render_execution_mermaid(execution, nodes, label)
    return "\n".join(out) + "\n"


def render_control_center_page(data, graph):
    """La pagina Wiki dello stato del progetto: quello che il Control Center risponde, senza server.

    Il Control Center e' una pagina locale — serve `python -m http.server` e un checkout. Chi legge
    la Wiki non ha ne' l'uno ne' l'altro, e fino a qui poteva sapere *quali feature esistono*
    (`Stato-delle-feature`) ma non **a che punto e' la consegna** ne' **cosa blocca cosa**.

    Non duplica `Stato-delle-feature`: quella elenca le feature per release e area, questa
    risponde alle domande che il Control Center aggiunge — i gate di release, la coda del lavoro
    umano, e l'execution graph. Ogni numero e' letto da `project-graph.json`, che e' generato:
    nessun conteggio nasce qui, e nessuna regola di stato viene riapplicata.
    """
    gates = graph.get("release_gates") or []
    green = sum(1 for g in gates if g.get("state") == "✅")
    queue = graph.get("editor_queue") or {}
    sessions = {s["id"]: s for s in (graph.get("editor_sessions") or [])}
    execution = graph.get("execution") or None
    diagnostics = graph.get("diagnostics") or {}
    counts = {}
    for entry in data["features"]:
        counts[entry["status"]] = counts.get(entry["status"], 0) + 1
    releases = {}
    for entry in data["features"]:
        releases[entry["release"]] = releases.get(entry["release"], 0) + 1

    out = [
        "# Stato del progetto",
        "",
        "> `GENERATO` · Vista d'insieme sopra gli artefatti del Feature Registry.",
        "> Non modificare a mano: si rigenera con "
        "`python scripts/feature_registry.py deploy --wiki-root <clone> --write`.",
        "> E' la stessa cosa che il **Project Control Center** mostra nel repository, per chi legge",
        "> la Wiki e non ha un checkout: la pagina locale e' interattiva, questa e' leggibile.",
        "",
        "## Quanto manca alla v0.1",
        "",
        f"**{green} gate di release verdi su {len(gates)}**. Il progresso si legge `N/M`, mai in",
        "percentuale: un gate `na` esce dal denominatore invece di gonfiare la frazione.",
        "",
        "| Gate | Cosa chiede | Stato |",
        "|---|---|:--:|",
    ]
    for gate in gates:
        request = (gate.get("request") or "—").replace("|", "\\|")
        out.append(f"| **{gate.get('id')}** | {request} | {gate.get('state') or '—'} |")

    out += [
        "",
        "## Le feature, in due tagli",
        "",
        f"**{data['count']} feature** · "
        + " · ".join(f"{rel} **{releases[rel]}**" for rel in RELEASE_ORDER if rel in releases)
        + ".",
        "",
        "| Stato | Quante |",
        "|---|--:|",
    ]
    for status in STATUS_ORDER:
        if counts.get(status):
            out.append(f"| `{status}` | {counts[status]} |")
    for status in sorted(s for s in counts if s in STATUS_OFF_SCALE):
        out.append(f"| `{status}` (fuori scala) | {counts[status]} |")
    out += [
        "",
        "L'elenco per release e area sta in [Stato delle feature](Stato-delle-feature).",
        "",
        "## Il lavoro che aspetta una persona",
        "",
        "Le sedute in editor, nell'ordine in cui la coda le mette. `ASSET` significa che l'uscita e'",
        "un asset da costruire e committare, `PIE` che e' un verdetto da dare guardando il gioco:",
        "e' il **tipo di lavoro**, non l'evidenza — `U7` e' `ASSET` e verifica anche due voci `PIE-*`.",
        "",
    ]
    labels = {"BLOCKING": "Blocca la v0.1 e si puo' fare adesso", "READY": "Si puo' fare",
              "WAITING": "Aspetta qualcosa", "DONE": "Chiuse"}
    for group in ("BLOCKING", "READY", "WAITING", "DONE"):
        ids = queue.get(group) or []
        if not ids:
            continue
        out += ["", f"### {group} — {labels[group]} ({len(ids)})", "",
                "| Seduta | Lane | Stato | Produce |", "|---|:--:|:--:|---|"]
        for sid in ids:
            session = sessions.get(sid) or {}
            produces = (session.get("produces") or "—").replace("|", "\\|")
            out.append(f"| **{sid}** {session.get('title') or ''} | "
                       f"`{(session.get('execution_lane') or '—').upper()}` | "
                       f"{session.get('state') or '—'} | {produces} |")

    if execution:
        stats = execution.get("stats") or {}
        by_readiness = stats.get("by_readiness") or {}
        out += [
            "",
            "## Cosa blocca cosa",
            "",
            f"La fetta «{execution.get('status')}» dell'execution graph: **{stats.get('nodes')} nodi**, "
            f"**{stats.get('hard_edges')} dipendenze dure**, "
            f"**{stats.get('soft_edges')} relazioni molli** (ordine consigliato o navigazione, che "
            "non bloccano).",
            "",
            "| Readiness | Quanti | Significa |",
            "|---|--:|---|",
        ]
        meaning = {
            "READY": "nessun prerequisito duro lo trattiene",
            "BLOCKED": "un prerequisito duro non e' risolto",
            "WAITING_FOR_PIE": "aspetta una seduta davanti all'editor",
            "WAITING_FOR_ASSET": "aspetta un asset da costruire",
            "DONE": "fatto",
            "UNKNOWN": "stato non misurabile senza rete",
        }
        for state in ("READY", "BLOCKED", "WAITING_FOR_PIE", "WAITING_FOR_ASSET", "DONE", "UNKNOWN"):
            if by_readiness.get(state):
                out.append(f"| `{state}` | {by_readiness[state]} | {meaning[state]} |")
        # Due numeri diversi che sembrano lo stesso, e vanno distinti a parole: `state_unknown`
        # conta i nodi **il cui stato** non e' leggibile offline; `by_readiness["UNKNOWN"]` conta
        # quelli la cui **readiness** non e' decidibile — cioe' quelli che *dipendono* da uno di
        # quelli. Il primo e' sempre >= il secondo, e affiancarli senza spiegarli produce due
        # cifre che il lettore non sa riconciliare (qui: 14 e 5).
        unknown = stats.get("state_unknown") or 0
        if unknown:
            blocked_by_ignorance = by_readiness.get("UNKNOWN", 0)
            out += [
                "",
                f"> ⚠️ Di **{unknown}** nodi su {stats.get('nodes')} non si conosce lo **stato**: il",
                "> generatore non parla con GitHub, e il checkpoint di quei nodi non dichiara il glifo",
                "> nell'owner. Non e' la stessa cosa della riga `UNKNOWN` qui sopra, che ne conta"
                f" **{blocked_by_ignorance}**: quella e' la *readiness* — i nodi che non si sa se",
                "> siano liberi perche' **dipendono** da uno di stato ignoto. Chi non ha prerequisiti",
                "> resta `READY` anche senza stato proprio, perche' nulla lo trattiene comunque.",
            ]

        nodes = {n["id"]: n for n in execution.get("nodes") or []}
        incoming, outgoing = {}, {}
        for edge in execution.get("edges") or []:
            if edge.get("hard"):
                incoming.setdefault(edge["to"], []).append(edge["from"])
                outgoing.setdefault(edge["from"], []).append(edge["to"])
        def label(nid):
            """Il nome corto di un endpoint. Una capability non e' un nodo del grafo — compare solo
            come estremo di un arco — quindi `nodes` non la conosce e senza questo ramo finiva in
            tabella come `capability:DecisionBoundary`, l'unico id grezzo fra venti etichette."""
            if str(nid).startswith("capability:"):
                return f"`{str(nid).split(':', 1)[1]}`"
            node = nodes.get(nid) or {}
            return f"#{node['ref']}" if node.get("kind") == "issue" else node.get("ref") or nid
        out += render_execution_mermaid(execution, nodes, label)
        out += [
            "",
            "### Le stesse dipendenze, in tabella",
            "",
            "| Nodo | Lane | Dominio | Readiness | Dipende da | Abilita |",
            "|---|:--:|---|:--:|---|---|",
        ]
        for node in execution.get("nodes") or []:
            up = ", ".join(label(x) for x in sorted(incoming.get(node["id"], []))) or "—"
            down = ", ".join(label(x) for x in sorted(outgoing.get(node["id"], []))) or "—"
            out.append(
                f"| **{label(node['id'])}** | `{(node.get('execution_lane') or '—').upper()}` | "
                f"{node.get('domain_group') or '—'} | `{node.get('readiness') or '—'}` | {up} | {down} |")
        out += [
            "",
            "Un `—` non e' un dato mancante: a sinistra significa che **niente lo trattiene**, a",
            "destra che **non sblocca niente di dichiarato**. Nella pagina locale gli stessi due casi",
            "si vedono come un collegamento staccato.",
        ]

        capabilities = execution.get("capabilities") or []
        if capabilities:
            out += [
                "",
                "### Capability",
                "",
                "Una capacita' che gli scenari chiedono e che il gioco non ha ancora. Finche' manca,",
                "gli scenari che la dichiarano restano `BLOCKED` — e questa tabella dice **chi la",
                "portera'**, che e' l'unica cosa che nessun'altra fonte sapeva.",
                "",
                "| Capability | Disponibile | La fornisce | Scenari in attesa |",
                "|---|:--:|---|--:|",
            ]
            for cap in capabilities:
                providers = ", ".join(label(p) for p in cap.get("providers") or []) or "—"
                out.append(f"| `{cap['id']}` | {'sì' if cap.get('available') else 'no'} | "
                           f"{providers} | {len(cap.get('scenario_consumers') or [])} |")

    errors = diagnostics.get("errors") or []
    warnings = diagnostics.get("warnings") or []
    out += [
        "",
        "## Diagnostica",
        "",
        f"**{len(errors)} errori** · **{len(warnings)} warning** dal validator del registry. "
        "Un errore fa fallire il gate e blocca la consegna; un warning e' una lacuna dichiarata.",
        "",
        "---",
        "",
        "Le viste di dettaglio vivono nel repository e non sulla Wiki, perche' citano file e righe:",
        "`docs/roadmap/*.shortlist.md` per epic, feature, scenari, milestone e sedute.",
    ]
    return "\n".join(out) + "\n"


def render_status_page(data):
    by_release = {rel: [] for rel in RELEASE_ORDER}
    for entry in data["features"]:
        by_release.setdefault(entry["release"], []).append(entry)

    out = [
        "# Stato delle feature",
        "",
        "> `GENERATO` · Vista del **Feature Registry** "
        "(`docs/roadmap/feature-registry.yaml` nel repository del gioco).",
        "> Non modificare a mano: si rigenera con "
        "`python scripts/feature_registry.py wiki`.",
        f"> Feature tracciate: **{data['count']}** · "
        f"Ultimo audit completo: **{(data['meta'].get('last_full_audit') or {}).get('date', '—')}** "
        f"su `{(data['meta'].get('last_full_audit') or {}).get('commit', '—')}`.",
        "",
        "Questa pagina dice **cosa esiste davvero**. Una meccanica descritta altrove nella Wiki e",
        "marcata qui come `SPECIFIED` o `DESIGNED` e' progettata, non giocabile: la Wiki racconta",
        "il gioco che sara', questa tabella dice a che punto e'.",
        "",
        "## Cosa significano gli stati",
        "",
        "| Stato | Significato |",
        "|---|---|",
        "| `IDEA` | Nominata, nessun documento che la definisca |",
        "| `DESIGNED` | Un brief la descrive, le regole non sono chiuse |",
        "| `SPECIFIED` | Regole decise e documentate, nessun codice |",
        "| `IMPLEMENTING` | Codice presente ma incompleto o non coperto da test |",
        "| `TESTABLE` | Implementata e coperta da test automatici |",
        "| `INTEGRATED` | Testata **e** dimostrata da uno scenario giocabile |",
        "| `RELEASE_READY` | Anche UI e documentazione allineate |",
        "| `DONE` | Verificata anche su build packaged |",
        "",
        "«Gate» conta i controlli superati sui controlli applicabili: `6/8` si verifica,",
        "«73%» no.",
        "",
    ]

    # Il titolo di `future` nomina l'ultima release pianificata, e scriverlo a mano l'aveva gia'
    # lasciato indietro: diceva «Oltre la v0.2» mentre `roadmap-post-v0.1.md` pianificava v0.3 e
    # v0.4. Derivarlo da `RELEASE_ORDER` fa si' che la prossima release lo aggiorni da sola.
    last_planned = [rel for rel in RELEASE_ORDER if rel != "future"][-1]
    titles = {rel: f"Release {rel}" for rel in RELEASE_ORDER if rel != "future"}
    titles["future"] = f"Oltre la {last_planned}"
    for release in RELEASE_ORDER:
        entries = by_release.get(release) or []
        if not entries:
            continue
        out += [f"## {titles.get(release, release)}", ""]
        for area in sorted({e["area"] for e in entries}):
            out += [
                f"### {area}",
                "",
                "| Feature | Titolo | Roadmap | Stato | Gate | Scenario |",
                "|---|---|---|---|---:|---|",
            ]
            for entry in sorted(entries, key=lambda e: e["feature_id"]):
                if entry["area"] != area:
                    continue
                ref = roadmap_ref(entry["roadmap"])
                if entry["scenarios"]:
                    scenario = "`" + "` · `".join(entry["scenarios"][:2]) + "`"
                elif entry["scenarios_planned"]:
                    scenario = "_pianificato_"
                else:
                    scenario = "—"
                out.append(
                    f"| `{entry['feature_id']}` | {entry['title']} | {ref} | "
                    f"**{entry['status']}** | {entry['gates_done']}/{entry['gates_applicable']} | "
                    f"{scenario} |"
                )
            out.append("")
    return "\n".join(out).rstrip() + "\n"


def apply_wiki_blocks(data, check=False):
    """Inserisce o aggiorna il blocco di stato nelle pagine referenziate dal registry."""
    by_page = {}
    for entry in data["features"]:
        for ref in entry["wiki_refs"]:
            if ref.startswith("wiki:"):
                continue
            by_page.setdefault(ref, []).append(entry)

    changed, stale = [], []
    for ref, entries in sorted(by_page.items()):
        path = os.path.join(REPO, ref)
        if not os.path.isfile(path):
            continue
        original = open(path, encoding="utf-8").read()
        text = original
        wanted = {e["feature_id"]: status_block(e) for e in entries}

        existing = set(re.findall(r"RT_FEATURE_STATUS:BEGIN ([A-Za-z0-9\-]+)", text))
        for fid in existing - set(wanted):
            stale.append(f"{ref}: blocco per {fid} non e' piu' referenziato dal registry")

        # I blocchi nuovi si inseriscono sempre in cima: iterando al contrario, l'ordine finale
        # sulla pagina e' quello del registry invece del suo rovescio.
        for fid, block in reversed(list(wanted.items())):
            pattern = re.compile(
                re.escape(MARKER_BEGIN.format(fid=fid)) + r".*?" + re.escape(MARKER_END.format(fid=fid)),
                re.S,
            )
            if pattern.search(text):
                text = pattern.sub(lambda _m, b=block: b, text)
            else:
                text = insert_block(text, block)

        if text != original:
            changed.append(ref)
            if not check:
                with open(path, "w", encoding="utf-8", newline="\n") as fh:
                    fh.write(text)
    return changed, stale


# Rinominare una pagina della GitHub Wiki ne cambia l'**URL pubblico**, e un link esterno o un
# segnalibro smette di funzionare anche per una questione di sole maiuscole: dove il clone ha gia'
# pubblicato un nome, vince il clone.
def deploy_name(source_ref, wiki_root=None):
    """Percorso della pagina nel clone della Wiki, che e' un repository separato.

    Cio' che e' piatto nella GitHub Wiki e' il **namespace degli URL**, non il filesystem: l'URL e'
    il solo nome del file, il percorso non ci entra. Quindi i nomi restano **globalmente unici**
    anche fra cartelle diverse, e i `[[link]]` sono per **nome**, mai per percorso. Le pagine indice
    stanno in root accanto a `Home`, perche' tre `index.md` in tre cartelle non possono convivere.

    **Dal 2026-08-10 (D-076) esistono due forme di ref, e non sono simmetriche.**

    `wiki:<PageName>` nomina una pagina che vive **solo** nel clone: `docs/wiki/` e' stata cancellata
    e il clone e' la fonte. Il nome non dice la cartella — e' voluto: sopravvive a una
    riorganizzazione delle cartelle, che e' esattamente cio' che e' successo il 2026-08-10 quando 89
    pagine sono uscite dalla root. Risolverlo richiede quindi di **guardare nel clone**, e senza
    `wiki_root` la funzione non puo' rispondere: restituisce `None` invece di tirare a indovinare.

    `docs/characters/...md` resta un percorso nel repository, perche' quelle pagine **restano** qui:
    sono il catalogo degli eroi, referenziato da cataloghi, ADR e test. Il clone ne e' solo il
    deploy.
    """
    ref = source_ref.replace("\\", "/")
    if ref.startswith("wiki:"):
        if not wiki_root:
            return None
        path = wiki_page_by_name(wiki_root, ref[len("wiki:"):])
        return os.path.relpath(path, wiki_root).replace("\\", "/") if path else None
    stem = os.path.splitext(os.path.basename(ref))[0]
    if ref.startswith("docs/characters/v0."):
        return "Personaggi/" + stem + ".md"
    if ref == "docs/characters/index.md":
        return "Personaggi.md"
    if ref == "docs/characters/paragon.md":
        return "Personaggi/paragon.md"
    return None


MANUAL_STATUS_LINE = re.compile(
    r"^> (?:[✅\U0001F7E1⚠️ ]*)?\*\*(?:Stato (?:v0\.1|nel gioco|repository)|Design status)"
    r"[:\*]",
)


def strip_manual_status(text):
    """Toglie i banner di stato scritti a mano: il blocco generato li sostituisce.

    Nella sorgente in-repo sono gia' spariti; nel clone della Wiki no, perche' e' un altro
    repository. Senza questo passaggio una pagina pubblicata mostrerebbe **due** stati, che e'
    peggio di zero: il lettore non sa quale dei due e' vecchio.
    """
    kept = [line for line in text.split("\n") if not MANUAL_STATUS_LINE.match(line)]
    return re.sub(r"\n{3,}", "\n\n", "\n".join(kept))


def wiki_pages(wiki_root):
    """I percorsi delle pagine presenti nel clone, per un confronto **case-sensitive** e ricorsivo.

    Case-sensitive perche' `os.path.isfile` su Windows non distingue le maiuscole: un nome di deploy
    sbagliato nel solo case passerebbe qui e fallirebbe su un filesystem case-sensitive, saltando in
    silenzio i blocchi di quella pagina.

    Ricorsivo perche' dal 2026-08-10 le pagine stanno in cartelle: un `os.listdir` non ricorsivo le
    dichiarerebbe tutte assenti in blocco.
    """
    found = set()
    for root, dirs, files in os.walk(wiki_root):
        dirs[:] = [d for d in dirs if d != ".git"]
        for name in files:
            if name.endswith(".md"):
                rel = os.path.relpath(os.path.join(root, name), wiki_root)
                found.add(rel.replace("\\", "/"))
    return found


def wiki_page_by_name(wiki_root, page_name):
    """Il percorso della pagina con questo **nome**, che nella Wiki e' l'unita' di indirizzamento.

    I riferimenti `wiki:<PageName>` nominano la pagina, non il file: dopo il passaggio alle cartelle
    non si puo' piu' assumere che stia in root.

    **L'iterazione e' ordinata, e non e' cosmesi.** `wiki_pages()` restituisce un `set`, il cui
    ordine in CPython non e' stabile fra run diversi. Finche' D-076 non e' atterrata questa funzione
    serviva solo a un controllo booleano di esistenza e il percorso restituito veniva scartato; ora
    decide **in quale file** scrivere il blocco di stato e come classificare una riga del workbook.
    Con due pagine omonime — che `check-docs-links.py --wiki-root` segnala come errore, perche' si
    contendono lo stesso URL — un ordine casuale farebbe scrivere in un file diverso a run diverse,
    senza che nulla lo dica. E' la stessa regola che il progetto applica a `TMap`/`TSet` in C++.
    """
    for rel in sorted(wiki_pages(wiki_root)):
        if os.path.splitext(os.path.basename(rel))[0] == page_name:
            return os.path.join(wiki_root, rel)
    return None


def apply_wiki_deploy(data, wiki_root, check=True):
    """Porta i blocchi generati nel clone della Wiki. Di norma si esegue con `--check`.

    Il clone e' un **altro repository**, pubblico, che altre sessioni possono avere in lavorazione:
    scriverci e' un'azione che si chiede, non si assume.
    """
    pages = wiki_pages(wiki_root)
    by_page = {}
    for entry in data["features"]:
        for ref in entry["wiki_refs"]:
            by_page.setdefault(ref, []).append(entry)

    changed, missing = [], []
    for ref, entries in sorted(by_page.items()):
        name = deploy_name(ref, wiki_root)
        if not name:
            missing.append(f"{ref}: nessuna pagina di deploy corrispondente")
            continue
        path = os.path.join(wiki_root, name)
        if name not in pages:
            near = next((p for p in pages if p.lower() == name.lower()), None)
            detail = f" (nel clone c'e' `{near}`: differisce solo nel case)" if near else ""
            missing.append(f"{ref} -> {name}: la pagina non esiste nel clone{detail}")
            continue
        original = open(path, encoding="utf-8").read()
        text = strip_manual_status(original)
        for entry in reversed(entries):
            fid = entry["feature_id"]
            block = status_block(entry)
            pattern = re.compile(
                re.escape(MARKER_BEGIN.format(fid=fid)) + r".*?" + re.escape(MARKER_END.format(fid=fid)),
                re.S,
            )
            text = pattern.sub(lambda _m, b=block: b, text) if pattern.search(text) else insert_block(text, block)
        if text != original:
            changed.append(name)
            if not check:
                with open(path, "w", encoding="utf-8", newline="\n") as fh:
                    fh.write(text)

    # La pagina di stato non ha piu' una sorgente nel repository. Fino a D-076 viveva in
    # `docs/wiki/feature-status.md`: il comando `wiki` la generava li', il deploy la **copiava** nel
    # clone, e fra i due passaggi c'era una finestra in cui la copia pubblicata era vecchia senza che
    # nulla lo dicesse. Ora si genera direttamente qui: una copia sola, nessuna finestra.
    # Due pagine generate direttamente qui, per la stessa ragione: una copia sola e nessuna
    # finestra in cui il pubblicato e' vecchio senza che nulla lo dica. `Stato-del-progetto` ha
    # bisogno anche del grafo — gate, coda, execution — e lo costruisce da se': e' l'unico punto
    # del deploy che lo richiede, e passarlo da fuori avrebbe obbligato ogni chiamante ad averlo.
    generated = [
        ("Stato-delle-feature.md", lambda: render_status_page(data)),
        ("Stato-del-progetto.md", lambda: render_control_center_page(data, build_graph(load_registry()))),
    ]
    for name, render in generated:
        target = os.path.join(wiki_root, name)
        content = render()
        current = open(target, encoding="utf-8").read() if os.path.isfile(target) else None
        if current != content:
            changed.append(name)
            if not check:
                with open(target, "w", encoding="utf-8", newline="\n") as fh:
                    fh.write(content)
    return changed, missing


def insert_block(text, block):
    """Il blocco va dopo il titolo e l'eventuale citazione introduttiva, prima del corpo."""
    lines = text.split("\n")
    index = 0
    for i, line in enumerate(lines):
        if line.startswith("# "):
            index = i + 1
            break
    while index < len(lines) and (lines[index].strip() == "" or lines[index].startswith(">")):
        index += 1
    prefix = lines[:index]
    suffix = lines[index:]
    if prefix and prefix[-1].strip() != "":
        prefix.append("")
    return "\n".join(prefix + [block, ""] + suffix)


ROADMAP_MARKER_BEGIN = "<!-- RT_FEATURE_BY_EPIC:BEGIN -->"
ROADMAP_MARKER_END = "<!-- RT_FEATURE_BY_EPIC:END -->"

SUITE_MARKER_BEGIN = "<!-- RT_SUITE_COUNT:BEGIN -->"
SUITE_MARKER_END = "<!-- RT_SUITE_COUNT:END -->"
SUITE_COUNT_TARGETS = (
    os.path.join(REPO, "docs", "README.md"),
    ROADMAP_V01,
)


def current_head():
    """Commit su cui la misura e' stata presa. Senza, il numero non e' verificabile."""
    head = os.path.join(REPO, ".git")
    try:
        import subprocess
        out = subprocess.run(["git", "rev-parse", "--short", "HEAD"], cwd=REPO,
                             capture_output=True, text=True, timeout=10)
        return out.stdout.strip() or "sconosciuto"
    except Exception:
        return "sconosciuto"


def suite_measure():
    """Conta i test unici e i file che li dichiarano. Stessa fonte del validator."""
    names = known_tests()
    files = 0
    if os.path.isdir(TESTS_DIR):
        pattern = re.compile(r'"RefactorTactics\.[A-Za-z0-9_.]+"')
        for entry in sorted(os.listdir(TESTS_DIR)):
            if not entry.endswith(".cpp"):
                continue
            with open(os.path.join(TESTS_DIR, entry), encoding="utf-8", errors="replace") as fh:
                if pattern.search(fh.read()):
                    files += 1
    return len(names), files


# Ripartizione della suite per area. I prefissi sono esaustivi **per costruzione**: se un test non
# rientra in nessuna categoria il generatore fallisce, invece di produrre una tabella che non somma.
# E' il difetto che questa tabella aveva: «somma esattamente a 456» era vero il giorno in cui fu
# scritta e mai piu' verificato.
SUITE_AREAS = [
    ("`Hex*` (mappa, path, vision, bot, blast, move, match)",
     ("Hex", "HexMap", "HexMapActor", "HexPath", "HexVision", "HexBot", "HexBotPlay",
      "HexBlast", "HexMove", "HexMatch"),
     "Coordinate, A\\*, LOS, bot, partita completa"),
    ("`Actions.*`", ("Actions",),
     "Ordine per priorità, permutazione-invarianza, fallback, mappatura di fase"),
    ("`Terrain.*` · `Status.*` · `Environment.*`", ("Terrain", "Status", "Environment"),
     "Superfici, stati temporanei, propagazione elettrica, fuoco/acqua"),
    ("`Combat.*` · `HexCombat.*`", ("Combat", "HexCombat"),
     "Danno dopo scudo, forme, LOS, niente fuoco amico"),
    ("`Reactions.*`", ("Reactions",),
     "Attivazione singola, trigger puro, reazioni componibili, privacy"),
    ("`HexSim.*`", ("HexSim",),
     "Snapshot, budget, collisioni simultanee, **replay divergence 0**"),
    ("`Match*` (allestimento, formato, fine partita)", ("Match", "MatchFormat", "MatchSetup"),
     "Le tre vie di fine partita e il `RoundLimit` da formato"),
    ("`Heroes.*`", ("Heroes",),
     "I 4 eroi corrispondono al catalogo, trade-off delle varianti"),
    ("`TurnLog.*`", ("TurnLog",),
     "Hash permutazione-invariante, serializzazione versionata, checksum"),
    ("`Scenario.*` · `ScenarioIndex.*`", ("Scenario", "ScenarioIndex"),
     "Harness: PASS/FAIL/ERROR/**BLOCKED**, identità e tag, niente bypass"),
    ("`Structures.*`", ("Structures",),
     "Porte come bordo (E9.3), ponti come arco (E9.4)"),
    ("`Playback.*` · `Preview.*` · `PlayerInput.*` · `ShowcaseRelay.*` · `Camera.*`",
     ("Playback", "Preview", "PlayerInput", "ShowcaseRelay", "Camera"),
     "Presentazione e input: non decidono, riproducono"),
    ("`Unit.*` · `Turn.*` · `Simulation.*` · `Movement.*`",
     ("Unit", "Turn", "Simulation", "Movement"),
     "Stato unità, **ciclo di vita dei piani**, determinismo del replay"),
    ("`Cover.*`", ("Cover",),
     "Copertura bassa e alta, bordi, danno a struttura e distruzione"),
    ("`Catalog.*`", ("Catalog",),
     "Invarianti del catalogo: solo interi, slot dichiarati, ID stabili"),
    ("`Pacing.*`", ("Pacing",), "Pacing del turno misurato"),
    ("`Perf.*`", ("Perf",), "Path mediana **0,025 ms** · resolver **0,41 ms/turno**"),
]


def run_log_filter(text):
    """Il filtro con cui la run e' stata lanciata, letto dal log.

    Serve perche' il confronto dichiarati/eseguiti ha senso solo **a parita' di perimetro**: una run
    lanciata con `RunTests RefactorTactics.HexMap` non deve risultare mancante di 600 test.
    UE scrive la riga `LogInit: Command Line:` con l'`-ExecCmds` completo, e i comandi si separano con
    `;` o `+` — entrambi visti nei log di questo repository.
    """
    match = re.search(r"Automation\s+RunTests\s+([^\";+\r\n]+)", text)
    if not match:
        return None
    return match.group(1).strip()


def run_log_executed(text):
    """I test che la run ha DAVVERO eseguito, dal log dell'automation."""
    return set(re.findall(r"Test Completed\..*?Path=\{([^}]+)\}", text))


def suite_run_diff(log_path):
    """Confronta i test DICHIARATI nei sorgenti con quelli ESEGUITI da una run.

    Il conteggio canonico del repository misura i nomi nei `.cpp`; il verde della CI riguarda quelli
    eseguiti, e i due possono divergere **senza che nulla lo dica**: e' il difetto di `#486`, misurato
    su un worktree ricostruito in modo incrementale, dove la suite riportava 632 e poi 633
    `Test Completed` con zero falliti su 634 dichiarati. I mancanti non erano `Fail` ne' `Skipped`:
    non comparivano affatto.

    Ritorna `(scope, executed, missing, extra)`. `missing` e' un ERRORE — un test che nessuno ha
    eseguito non e' ne' rosso ne' verde, e' invisibile. `extra` e' un avviso: un test eseguito che i
    sorgenti non dichiarano **dove il contatore guarda** (`Source/RefactorTactics/Tests/`), quindi
    dice che il perimetro del contatore e' piu' stretto della realta'.
    """
    with open(log_path, encoding="utf-8", errors="replace") as fh:
        text = fh.read()

    executed = run_log_executed(text)
    prefix = run_log_filter(text) or "RefactorTactics"
    # Il filtro di UE e' per prefisso: `RunTests RefactorTactics.HexMap` prende anche
    # `RefactorTactics.HexMapActor`. Riprodurlo qui, o il gate accuserebbe assenze che non esistono.
    scope = {name for name in known_tests() if name.startswith(prefix)}

    return scope, executed, sorted(scope - executed), sorted(executed - scope)


def suite_breakdown():
    """Conteggio per area. Solleva se un test non rientra in nessuna categoria dichiarata."""
    names = known_tests()
    buckets = {label: 0 for label, _p, _d in SUITE_AREAS}
    prefix_to_label = {}
    for label, prefixes, _desc in SUITE_AREAS:
        for prefix in prefixes:
            prefix_to_label[prefix] = label

    orphans = []
    for name in names:
        parts = name.split(".")
        area = parts[1] if len(parts) > 1 else ""
        label = prefix_to_label.get(area)
        if label is None:
            orphans.append(name)
        else:
            buckets[label] += 1
    return buckets, sorted(orphans)


def render_suite_count():
    """Il conteggio della suite come valore GENERATO.

    Era scritto a mano in due posti e ha divertito cinque volte, l'ultima di 34 test. Il comando di
    misura era gia' nel documento: se il comando esiste, il numero non si copia.

    Restano fuori i numeri STORICI nelle spec di checkpoint (362, 359, 347...): fotografano cosa era
    vero alla chiusura di quel CP e non devono aggiornarsi mai. Non hanno marker, il generatore non
    li vede.
    """
    count, files = suite_measure()
    buckets, orphans = suite_breakdown()
    lines = [
        SUITE_MARKER_BEGIN,
        f"**{count} test unici in {files} file** — misurati su `{current_head()}`.",
        "",
        "Generato da `python scripts/feature_registry.py suite`: **non si aggiorna a mano**. Era "
        "scritto a mano in due documenti ed è divergito cinque volte.",
        "",
        "| Area | Test | Cosa fissa |",
        "|---|---:|---|",
    ]
    for label, _prefixes, desc in SUITE_AREAS:
        lines.append(f"| {label} | {buckets[label]} | {desc} |")
    lines.append(f"| **totale** | **{sum(buckets.values())}** | |")
    if orphans:
        lines += ["", "> ⚠️ Test fuori da ogni area dichiarata: " +
                  ", ".join(f"`{o}`" for o in orphans[:8])]
    lines.append(SUITE_MARKER_END)
    return "\n".join(lines)


SUITE_MEASURED_ON = re.compile(r"— misurati su `[^`]*`\.")


def suite_substance(block):
    """Il blocco senza il commit su cui la misura e' stata presa.

    `--check` deve confrontare la **sostanza** — conteggio, file, ripartizione per area — e non il
    commit corrente, altrimenti il gate e' rosso ovunque tranne che sul commit esatto in cui e' stato
    rigenerato: cioe' mai su `main`, perche' un merge sposta `HEAD` (`#488`).

    Un allarme che suona sempre non lo guarda piu' nessuno, ed e' il motivo per cui il conteggio ha
    potuto restare fermo a 573 test mentre erano 634: lo strumento *stava gia' segnalando*,
    indistinguibile dal rumore che produceva a ogni commit.

    La riga `misurati su ...` **resta scritta**: e' l'unica cosa che rende il numero verificabile. Non
    e' confrontata, ed e' corretto — quando la sostanza non cambia, il commit registrato indica pur
    sempre un albero in cui quel numero era vero.
    """
    return SUITE_MEASURED_ON.sub("— misurati su `<commit>`.", block)


def apply_suite_count(check=False):
    block = render_suite_count()
    pattern = re.compile(
        re.escape(SUITE_MARKER_BEGIN) + r".*?" + re.escape(SUITE_MARKER_END), re.S)
    changed = []
    for path in SUITE_COUNT_TARGETS:
        if not os.path.isfile(path):
            continue
        original = open(path, encoding="utf-8").read()
        if SUITE_MARKER_BEGIN not in original:
            continue
        found = pattern.search(original)
        # Confronto sulla sostanza anche in SCRITTURA, non solo con `--check`: senza, `suite`
        # produrrebbe un diff su ogni branch — rumore in ogni PR che lo tocca — e il commit
        # registrato inseguirebbe `HEAD` invece di dire dove la misura e' stata presa.
        if found and suite_substance(found.group(0)) == suite_substance(block):
            continue
        text = pattern.sub(lambda _m: block, original)
        if text != original:
            changed.append(os.path.relpath(path, REPO).replace("\\", "/"))
            if not check:
                with open(path, "w", encoding="utf-8", newline="\n") as fh:
                    fh.write(text)
    return changed


def render_features_by_epic(data):
    """Mappa epic → feature per la roadmap di release: generata, non ricopiata a mano."""
    by_epic = {}
    unassigned = []
    out_of_scope = []
    for entry in data["features"]:
        epic = entry["roadmap"].get("epic")
        if epic:
            by_epic.setdefault(epic, []).append(entry)
        elif entry["release"] == "v0.1":
            # Un buco e una decisione non si mescolano: la prima tabella deve poter restare vuota.
            if entry["roadmap"].get("out_of_release_scope"):
                out_of_scope.append(entry)
            else:
                unassigned.append(entry)

    lines = [
        ROADMAP_MARKER_BEGIN,
        "",
        "| Epic | Feature | Stato | Gate |",
        "|---|---|---|---:|",
    ]
    for epic in sorted(by_epic, key=lambda e: int(e[1:])):
        entries = sorted(by_epic[epic], key=lambda e: e["feature_id"])
        for i, entry in enumerate(entries):
            label = f"**{epic}**" if i == 0 else ""
            title = entry["title"]
            if entry.get("completed_by"):
                # Il rimando sta nella riga, non in una nota: e' cio' che impedisce di contare
                # due volte la stessa mancanza.
                title += " — completata da " + ", ".join(f"`{c}`" for c in entry["completed_by"])
            lines.append(
                f"| {label} | `{entry['feature_id']}` — {title} | "
                f"{entry['status']} | {entry['gates_done']}/{entry['gates_applicable']} |"
            )
    lines.append("")
    if unassigned:
        lines += [
            "> ⚠️ **Feature v0.1 senza assegnazione** — lavoro dentro lo scope della release che nessuna",
            "> epic copre e che nessuno ha dichiarato fuori scope. Questa tabella **deve restare vuota**:",
            "> se compare una riga, o le si assegna un'epic o si dichiara `out_of_release_scope` con un",
            "> motivo.",
            "",
            "| Feature | Vista | Stato | Gate |",
            "|---|---|---|---:|",
        ]
        for entry in sorted(unassigned, key=lambda e: e["feature_id"]):
            where = entry["roadmap"].get("milestone") or "—"
            lines.append(
                f"| `{entry['feature_id']}` — {entry['title']} | {where} | "
                f"{entry['status']} | {entry['gates_done']}/{entry['gates_applicable']} |"
            )
        lines.append("")
    if out_of_scope:
        lines += [
            "**Fuori dalla vista di release, per decisione** — esiste, è tracciato, ma non è contenuto",
            "della v0.1. Dichiararlo è tracciabilità quanto assegnare un'epic: quello che non va bene è il",
            "silenzio.",
            "",
            "| Feature | Vista | Perché fuori scope |",
            "|---|---|---|",
        ]
        for entry in sorted(out_of_scope, key=lambda e: e["feature_id"]):
            where = entry["roadmap"].get("milestone") or "—"
            reason = " ".join(entry["roadmap"]["out_of_release_scope"].split())
            lines.append(f"| `{entry['feature_id']}` — {entry['title']} | {where} | {reason} |")
        lines.append("")
    lines.append(ROADMAP_MARKER_END)
    return "\n".join(lines)


def apply_roadmap_block(data, check=False):
    if not os.path.isfile(ROADMAP_V01):
        return False
    original = open(ROADMAP_V01, encoding="utf-8").read()
    if ROADMAP_MARKER_BEGIN not in original:
        return False
    pattern = re.compile(
        re.escape(ROADMAP_MARKER_BEGIN) + r".*?" + re.escape(ROADMAP_MARKER_END), re.S)
    block = render_features_by_epic(data)
    text = pattern.sub(lambda _m: block, original)
    if text == original:
        return False
    if not check:
        with open(ROADMAP_V01, "w", encoding="utf-8", newline="\n") as fh:
            fh.write(text)
    return True


CHARACTERS_WORKBOOK = os.path.join(
    REPO, "docs", "characters", "data", "RefactorTactics_Characters_Wiki_Data_v0.4.xlsx")
REFS_SHEET = "15_Wiki_Feature_Refs"


def wiki_entity_id(ref, wiki_root=None):
    """Entita' della Wiki a cui una pagina corrisponde, per il workbook.

    La famiglia dell'entita' (`Faction` / `Mechanic` / `Guide`) e' la **cartella** che contiene la
    pagina, non il suo nome. Fino a D-076 la cartella stava nel ref (`docs/wiki/meccaniche/...`);
    ora i ref sono `wiki:<PageName>` e la cartella vive solo nel clone, quindi serve `wiki_root`.

    Senza, restituisce `None` e la relazione sparisce dal foglio. Il chiamante deve accorgersene e
    fermarsi: un workbook con **meno righe** e nessun errore e' il modo in cui una copertura finta
    passa inosservata.
    """
    ref = ref.replace("\\", "/")
    stem = os.path.splitext(os.path.basename(ref))[0]
    if ref.startswith("docs/characters/v0.") and stem != "index":
        return "Hero." + stem.capitalize()
    if ref.startswith("wiki:"):
        page = ref[len("wiki:"):]
        path = wiki_page_by_name(wiki_root, page) if wiki_root else None
        if not path:
            return None
        rel = os.path.relpath(path, wiki_root).replace("\\", "/")
        folder = rel.split("/")[0] if "/" in rel else ""
        stem = os.path.splitext(os.path.basename(rel))[0]
        prefix = {"Fazioni": "Faction.", "Meccaniche": "Mechanic.", "Guida": "Guide."}.get(folder)
        if not prefix:
            return None
        return prefix + "".join(p.capitalize() for p in stem.split("-"))
    return None


def feature_refs_rows(data, wiki_root=None):
    """Relazioni entita' Wiki → FeatureId, derivate dal registry. Nessuno stato: solo riferimenti.

    Restituisce anche i ref `wiki:` che **non** si sono risolti. Ignorarli produrrebbe un foglio piu'
    corto senza un errore: la firma li rende visibili al chiamante, che decide se fermarsi.
    """
    rows, unresolved = [], []
    for entry in data["features"]:
        for ref in entry["wiki_refs"]:
            # «Non e' un'entita'» e «non si risolve» sono due cose diverse, e confonderle costa in
            # entrambi i versi. Le pagine indice (`Fazioni`, `Meccaniche`) stanno in root del clone
            # e non sono entita' del workbook — non lo erano nemmeno prima di D-076. Una pagina che
            # invece **non esiste** nel clone e' un ref rotto, e va detto.
            if ref.startswith("wiki:") and not (wiki_root and wiki_page_by_name(wiki_root, ref[len("wiki:"):])):
                unresolved.append(ref)
                continue
            entity = wiki_entity_id(ref, wiki_root)
            if not entity:
                continue
            if entity.startswith("Faction."):
                relation = "belongs-to"
            elif entry["area"] == "Characters":
                relation = "release"
            elif entity.startswith("Hero."):
                relation = "demonstrates"
            else:
                relation = "documents"
            rows.append((entity, entry["feature_id"], relation, ref))
        for sid in entry["scenarios_planned"]:
            if sid.startswith("Team."):
                rows.append(("Scenario." + sid, entry["feature_id"], "validates", "pianificato"))
    return sorted(set(rows)), sorted(set(unresolved))


def apply_workbook(data, check=False, wiki_root=None):
    """Riscrive la sheet delle relazioni nel workbook di character authoring.

    Il workbook **non** tiene lo stato: quello vive nel registry. Qui stanno solo i riferimenti,
    perche' una riga «Design_Status = IMPLEMENTED» scritta a mano e' la terza copia dello stesso
    fatto, e la terza copia e' quella che nessuno ricorda di aggiornare.
    """
    try:
        import openpyxl
    except ImportError:
        return None, "serve openpyxl: pip install openpyxl"
    if not os.path.isfile(CHARACTERS_WORKBOOK):
        return None, f"workbook non trovato: {CHARACTERS_WORKBOOK}"

    rows, unresolved = feature_refs_rows(data, wiki_root)
    # Le pagine di gioco vivono nel clone (D-076): senza `--wiki-root` non si risolvono, e il foglio
    # perderebbe in silenzio le relazioni `Guide.*`, `Mechanic.*` e `Faction.*` — le tre famiglie su
    # quattro. Meglio non scrivere nulla che scrivere una copertura dimezzata.
    if unresolved:
        return None, (
            f"{len(unresolved)} riferimenti `wiki:` non risolti"
            + ("" if wiki_root else " (manca --wiki-root)")
            + ": " + ", ".join(unresolved[:5]) + ("…" if len(unresolved) > 5 else "")
        )
    workbook = openpyxl.load_workbook(CHARACTERS_WORKBOOK)
    existing = []
    if REFS_SHEET in workbook.sheetnames:
        sheet = workbook[REFS_SHEET]
        for row in sheet.iter_rows(min_row=4, values_only=True):
            if row and row[0]:
                existing.append(tuple(str(c) if c is not None else "" for c in row[:4]))
    if existing == [tuple(r) for r in rows]:
        return [], None
    if check:
        return rows, None

    if REFS_SHEET in workbook.sheetnames:
        del workbook[REFS_SHEET]
    sheet = workbook.create_sheet(REFS_SHEET)
    sheet["A1"] = ("Riferimenti Wiki -> Feature Registry - GENERATO, non modificare a mano "
                   "(python scripts/feature_registry.py workbook)")
    sheet["A2"] = ("Lo stato di una feature NON vive qui: sta in docs/roadmap/feature-registry.yaml "
                   "ed e' derivato dai gate. Qui ci sono solo i riferimenti.")
    for column, header in enumerate(("WikiEntityId", "FeatureId", "Relation", "Source"), start=1):
        sheet.cell(row=3, column=column, value=header)
    for index, row in enumerate(rows, start=4):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=index, column=column, value=value)
    for column, width in zip("ABCD", (34, 38, 14, 46)):
        sheet.column_dimensions[column].width = width
    workbook.save(CHARACTERS_WORKBOOK)
    return rows, None


def render_audit(data):
    out = [
        "| FeatureId | Titolo | Release | Stato | Roadmap | Owner spec | Test | Scenari | Wiki |",
        "|---|---|---|---|---|---:|---:|---:|---:|",
    ]
    for entry in sorted(data["features"], key=lambda e: (e["release"], e["area"], e["feature_id"])):
        ref = roadmap_ref(entry["roadmap"])
        out.append(
            f"| `{entry['feature_id']}` | {entry['title']} | {entry['release']} | "
            f"{entry['status']} ({entry['gates_done']}/{entry['gates_applicable']}) | {ref} | "
            f"{len(entry['owner_specs'])} | {len(entry['tests'])} | "
            f"{len(entry['scenarios'])} | {len(entry['wiki_refs'])} |"
        )
    return "\n".join(out)


# ---------------------------------------------------------------------------
# Shortlist — le quattro viste corte
#
# Una shortlist e' una vista corta di qualcosa che vive altrove: se ne ricopia lo stato a mano,
# diverge come e' gia' successo quattro volte col conteggio dei test. Qui ogni numero e ogni
# simbolo di stato arrivano dalla propria sorgente, misurati:
#
#   epic       -> `roadmap-v0.1.md` §2.1        (la sola vista dello stato delle epic)
#   milestone  -> `roadmap-checkpoint.md`       (vista di esecuzione)
#   gate       -> `v0.1-definition-of-done.md`  (§3, G1-G15)
#   feature    -> `feature-registry.yaml`       (stato derivato dai gate)
#   scenari    -> `Scenarios/` + `RTScenarioSession.cpp` (capability disponibili, dal CODICE)
#
# Cio' che NON e' derivabile — la riga di descrizione scritta da una persona — viene
# **preservata** dal file esistente e reimpaginata. Una feature nuova compare con `—`: e' un
# buco visibile, non una riga che sparisce.
# ---------------------------------------------------------------------------

SHORTLIST_DIR = os.path.join(REPO, "docs", "roadmap")
DOD_V01 = os.path.join(REPO, "docs", "roadmap", "v0.1-definition-of-done.md")
SCENARIO_SESSION_CPP = os.path.join(
    REPO, "Source", "RefactorTactics", "ScenarioHarness", "RTScenarioSession.cpp")
EDITOR_SESSIONS_YAML = os.path.join(SHORTLIST_DIR, "editor-sessions.yaml")
PIE_REGISTRY = os.path.join(REPO, "docs", "technical", "test-manuali-pie.md")

SHORTLIST_MARKERS = {
    "epics": "RT_SHORTLIST_EPICS",
    "releases": "RT_SHORTLIST_RELEASES",
    "features": "RT_SHORTLIST_FEATURES",
    "scenarios": "RT_SHORTLIST_SCENARIOS",
    "milestones": "RT_SHORTLIST_MILESTONES",
    "gates": "RT_SHORTLIST_MILESTONES_GATES",
    "editor": "RT_SHORTLIST_EDITOR",
}
SHORTLIST_FILES = {
    "epics": os.path.join(SHORTLIST_DIR, "roadmap.shortlist.md"),
    "releases": os.path.join(SHORTLIST_DIR, "roadmap.shortlist.md"),
    "features": os.path.join(SHORTLIST_DIR, "featuremap.shortlist.md"),
    "scenarios": os.path.join(SHORTLIST_DIR, "scenariomap.shortlist.md"),
    "milestones": os.path.join(SHORTLIST_DIR, "milestonemap.shortlist.md"),
    "gates": os.path.join(SHORTLIST_DIR, "milestonemap.shortlist.md"),
    "editor": os.path.join(SHORTLIST_DIR, "editormap.shortlist.md"),
}

# I simboli di stato in uso nei documenti di roadmap. `⌫` dichiara «rimosso dal repo».
STATUS_GLYPHS = "✅🟡⏳⌫"


def epic_status():
    """Stato per epic, letto da `roadmap-v0.1.md` §2.1 — che ne e' l'unico owner.

    Non si deriva dai gate delle feature: sarebbe una seconda regola, e due regole sullo stesso
    fatto sono il difetto che questo script esiste per non ripetere. Un'epic senza riga nella
    tabella resta senza stato, che e' un'informazione: significa che l'owner non la copre.
    """
    if not os.path.isfile(ROADMAP_V01):
        return {}
    text = open(ROADMAP_V01, encoding="utf-8").read()
    found = re.findall(r"^\| \*\*(E\d+)\*\*[^|]*\|\s*([" + STATUS_GLYPHS + r"])", text, re.M)
    return {epic: glyph for epic, glyph in found}


def milestone_status():
    """Stato per milestone da `roadmap-checkpoint.md`, con le divergenze **dichiarate**.

    Il file contiene due tabelle che parlano delle stesse milestone (`Stato attuale` e `Effetto
    delle epic sulle milestone`). Quando divergono, la shortlist non sceglie in silenzio: riporta
    la lettura piu' conservativa e nomina lo scarto, perche' la correzione va fatta nell'owner.
    """
    if not os.path.isfile(ROADMAP_CHECKPOINT):
        return {}, []
    text = open(ROADMAP_CHECKPOINT, encoding="utf-8").read()
    found = re.findall(r"^\| \*\*(M\d+)\*\*[^|]*\|\s*([" + STATUS_GLYPHS + r"])", text, re.M)
    seen, conflicts = {}, []
    rank = {"⏳": 0, "🟡": 1, "✅": 2, "⌫": 3}
    for milestone, glyph in found:
        if milestone in seen and seen[milestone] != glyph:
            conflicts.append((milestone, seen[milestone], glyph))
            # Conservativo: fra due letture vince quella che dichiara meno lavoro finito.
            if rank.get(glyph, 0) < rank.get(seen[milestone], 0):
                seen[milestone] = glyph
        else:
            seen.setdefault(milestone, glyph)
    return seen, conflicts


def post_v01_epics():
    """Le epic oltre la v0.1, da `roadmap-post-v0.1.md` — owner dichiarato delle release v0.2-v0.4.

    Tre letture dallo stesso file, perche' tre fatti diversi vivono in tre posti diversi:

    - la **release** dalla tabella «Le release». E' la sola riga che decide dove sta un'epic, e la
      colonna mescola intervalli (`E22-E26`) ed elenchi (`**E35** · **E36**`): entrambi vanno
      espansi, perche' un'epic dentro un intervallo non e' meno assegnata di una nominata.
    - **titolo e priorita'** dall'heading della propria sezione.
    - la **issue** dalla riga `**Tracciata su GitHub**`.

    Fino al 2026-08-13 questa vista era ricopiata a mano nella §3 di `roadmap.shortlist.md`, che si
    dichiarava «non generata». Aveva gia' perso **E36**, **E38** ed **E39** — tre epic che la
    tabella owner elenca in v0.2 — e mostrava **E25** senza issue benche' `#265` esista dal
    2026-08-08 e il documento owner la citi. Non erano sviste: una tabella copiata diverge dal suo
    owner appena l'owner cambia, e nessun gate confrontava le due.

    Un'epic con una sezione ma senza riga nella tabella resta **senza release**, e non e' un buco
    del parser: e' il caso di `E37`. La sua issue `#555` si intitola `[EPIC v0.4]`, la tabella owner
    non la assegna a nessuna release, e le due fonti non si conciliano indovinando — tanto meno
    dalla posizione del testo, che la mette dopo il titolo «v0.4» per ragioni di impaginazione.
    La vista dichiara lo scarto e lo lascia a chi possiede la tabella.
    """
    if not os.path.isfile(ROADMAP_POST_V01):
        return []
    text = open(ROADMAP_POST_V01, encoding="utf-8").read()

    release_of = {}
    for release, cell in re.findall(r"^\|\s*\*\*(v0\.\d)\*\*\s*\|[^|]*\|[^|]*\|([^|]*)\|", text, re.M):
        epics = []
        for first, last in re.findall(r"E(\d+)\s*[–—-]\s*E(\d+)", cell):
            epics += [f"E{n}" for n in range(int(first), int(last) + 1)]
        # Gli intervalli si tolgono prima di raccogliere i singoli, o `E22-E26` conterebbe anche
        # i suoi due estremi una seconda volta.
        singles = re.sub(r"E\d+\s*[–—-]\s*E\d+", " ", cell)
        epics += [f"E{n}" for n in re.findall(r"E(\d+)", singles)]
        for epic in epics:
            release_of[epic] = release

    sections = list(re.finditer(r"^### (E\d+) — (.+?) · (P\d)\s*$", text, re.M))
    out = []
    for index, match in enumerate(sections):
        end = sections[index + 1].start() if index + 1 < len(sections) else len(text)
        body = text[match.start():end]
        issue = re.search(r"\*\*Tracciata su GitHub\*\*[^\n]*?epic \[#(\d+)\]", body)
        out.append({
            "epic": match.group(1),
            "title": match.group(2).strip(),
            "priority": match.group(3),
            "release": release_of.get(match.group(1)),
            "issue": int(issue.group(1)) if issue else None,
        })
    return out


def release_gates():
    """I gate `G1`-`G15` dalla Definition of Done: id, richiesta e stato dichiarato."""
    if not os.path.isfile(DOD_V01):
        return []
    text = open(DOD_V01, encoding="utf-8").read()
    gates = []
    for row in re.findall(r"^\| \*\*(G\d+)\*\* \|(.+)$", text, re.M):
        gate_id, rest = row
        cells = [c.strip() for c in rest.split("|")]
        # `| id | richiesta | evidenza | stato |` — l'ultima cella piena e' lo stato.
        cells = [c for c in cells if c]
        request = cells[0] if cells else ""
        state = ""
        for cell in reversed(cells):
            if any(g in cell for g in STATUS_GLYPHS):
                state = cell
                break
        gates.append((gate_id, request, state))
    return gates


def available_capabilities():
    """Le capability che il gioco possiede oggi, lette dal CODICE.

    L'elenco sta in `RTScenarioSession.cpp` e non nei dati proprio perche' un JSON che dichiari
    disponibile una capability inesistente produrrebbe il primo verde bugiardo. Questa funzione
    lo misura li' per la stessa ragione: qualunque copia in un documento invecchia.
    """
    if not os.path.isfile(SCENARIO_SESSION_CPP):
        return set()
    text = open(SCENARIO_SESSION_CPP, encoding="utf-8", errors="replace").read()
    block = re.search(r"static const TSet<FString> Available\s*=\s*\{(.*?)\};", text, re.S)
    if not block:
        return set()
    return set(re.findall(r'TEXT\("([A-Za-z0-9_]+)"\)', block.group(1)))


def scenario_corpus():
    """Ogni scenario versionato con le capability che chiede.

    `requires` puo' comparire a piu' livelli (scenario o singolo turno): si raccolgono tutti,
    perche' basta un turno bloccato perche' lo scenario esca `BLOCKED`.
    """
    corpus = []
    for root, _dirs, files in os.walk(SCENARIOS_DIR):
        for name in sorted(files):
            if not name.endswith(".json") or name.startswith("_"):
                continue
            path = os.path.join(root, name)
            try:
                with open(path, encoding="utf-8-sig") as fh:
                    data = json.load(fh)
            except (json.JSONDecodeError, OSError):
                continue
            sid = data.get("scenarioId")
            if not sid:
                continue
            requires = set()

            def collect(node):
                if isinstance(node, dict):
                    for key, value in node.items():
                        if key == "requires" and isinstance(value, list):
                            requires.update(str(v) for v in value)
                        else:
                            collect(value)
                elif isinstance(node, list):
                    for item in node:
                        collect(item)

            collect(data)
            rel_path = os.path.relpath(path, REPO).replace("\\", "/")
            corpus.append({"id": sid, "requires": sorted(requires), "path": rel_path})
    return sorted(corpus, key=lambda s: s["id"])


def preserved_column(path, marker, key_pattern):
    """Le descrizioni scritte a mano nell'ultima colonna del blocco precedente.

    Il generatore riscrive le colonne derivate e **restituisce** questa: e' l'unica parte che una
    macchina non sa produrre, e cancellarla a ogni run renderebbe il comando inutilizzabile.
    """
    if not os.path.isfile(path):
        return {}
    text = open(path, encoding="utf-8").read()
    block = re.search(
        re.escape(f"<!-- {marker}:BEGIN -->") + r"(.*?)" + re.escape(f"<!-- {marker}:END -->"),
        text, re.S)
    if not block:
        return {}
    kept = {}
    for line in block.group(1).split("\n"):
        if not line.startswith("|") or line.startswith("|---"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) < 2:
            continue
        key = None
        for cell in cells:
            match = re.search(key_pattern, cell)
            if match:
                key = match.group(1)
                break
        if key and cells[-1] and cells[-1] != "—":
            kept[key] = cells[-1]
    return kept


def wrap_block(marker, lines):
    return "\n".join([f"<!-- {marker}:BEGIN -->", ""] + lines + ["", f"<!-- {marker}:END -->"])


def render_shortlist_epics(data):
    marker = SHORTLIST_MARKERS["epics"]
    kept = preserved_column(SHORTLIST_FILES["epics"], marker, r"\*\*(E\d+)\*\*")
    statuses = epic_status()
    by_epic = {}
    for entry in data["features"]:
        epic = entry["roadmap"].get("epic")
        if epic:
            by_epic.setdefault(epic, []).append(entry)
    epics = sorted(set(statuses) | set(by_epic), key=lambda e: int(e[1:]))
    lines = [
        "| Epic | Stato | Feature | Gate | In una riga |",
        "|---|:--:|--:|--:|---|",
    ]
    for epic in epics:
        entries = by_epic.get(epic, [])
        done = sum(e["gates_done"] for e in entries)
        total = sum(e["gates_applicable"] for e in entries)
        gates = f"{done}/{total}" if total else "—"
        lines.append(
            f"| **{epic}** | {statuses.get(epic, '—')} | {len(entries) or '—'} | {gates} | "
            f"{kept.get(epic, '—')} |"
        )
    missing = [e for e in epics if e not in statuses]
    lines.append("")
    lines.append(
        f"**{len(epics)} epic** · stato da [`roadmap-v0.1.md`](roadmap-v0.1.md) §2.1 · "
        f"gate dal Feature Registry.")
    if missing:
        lines += [
            "",
            "> ⚠️ **Senza stato dichiarato nell'owner**: " + ", ".join(f"**{e}**" for e in missing) +
            ". §2.1 non le copre — non è una svista di questa vista, è una riga che manca là.",
        ]
    return wrap_block(marker, lines)


def render_shortlist_releases(data):
    """§3 di `roadmap.shortlist.md`: gli step oltre la v0.1, generati dal loro owner.

    La colonna *In una riga* resta umana e viene conservata, come nelle altre shortlist. Tutto il
    resto — release, titolo, priorita', issue — arriva da `roadmap-post-v0.1.md`, cosi' che
    aggiungere un'epic la' la faccia comparire qui senza che nessuno se ne ricordi.

    Le Feature per epic si contano dal registry: e' la riga che dice se un'epic pianificata ha gia'
    una capacita' dichiarata o e' ancora solo un titolo, ed e' proprio la domanda che la §3
    manuale non sapeva rispondere.
    """
    marker = SHORTLIST_MARKERS["releases"]
    kept = preserved_column(SHORTLIST_FILES["releases"], marker, r"\*\*(E\d+)\*\*")
    epics = [e for e in post_v01_epics() if e["release"] != "v0.1"]

    by_epic = {}
    for entry in data["features"]:
        epic = entry["roadmap"].get("epic")
        if epic:
            by_epic.setdefault(epic, []).append(entry)

    lines = [
        "| Step | Rel. | Issue | Feature | In una riga |",
        "|---|:--:|---|--:|---|",
    ]
    unassigned = []
    # L'ordine delle release e' `RELEASE_ORDER`, non quello alfabetico: era l'ultima delle quattro
    # copie che questa stessa consegna esiste per eliminare, e sarebbe bastata una `v0.10` — o una
    # release con un nome invece che un numero — per farla divergere in silenzio dalle altre viste.
    rel_rank = {rel: i for i, rel in enumerate(RELEASE_ORDER)}
    def release_key(rel):
        return rel_rank.get(rel, len(RELEASE_ORDER))
    for entry in sorted(epics, key=lambda e: (release_key(e["release"]), int(e["epic"][1:]))):
        epic = entry["epic"]
        if entry["release"] is None:
            unassigned.append(entry)
            continue
        issue = f"`#{entry['issue']}`" if entry["issue"] else "—"
        count = len(by_epic.get(epic, []))
        lines.append(
            f"| **{epic}** {entry['title']} · {entry['priority']} | {entry['release']} | {issue} | "
            f"{count or '—'} | {kept.get(epic, '—')} |")

    releases = sorted({e["release"] for e in epics if e["release"]}, key=release_key)
    lines += [
        "",
        f"**{len([e for e in epics if e['release']])} epic** su {len(releases)} release "
        f"({' · '.join(releases)}) · release, titolo e issue da "
        f"[`roadmap-post-v0.1.md`](roadmap-post-v0.1.md) · Feature dal Feature Registry.",
    ]

    # Perche' la colonna Feature e' quasi tutta `—`, e perche' dirlo invece di nasconderlo: il
    # registry dichiara `roadmap.epic` quasi solo per E1-E21. Fra le 21 feature post-v0.1 **una**
    # lo dichiara (`RT-FEAT-REPLAY-ARCHIVE` -> `E12`, un'epic della v0.1 che la sua estensione
    # v0.2 continua a citare); le altre 20 no, e senza quel campo la §3 non puo' rispondere a
    # «quali capacita' porta questa epic».
    #
    # ⚠️ Il conteggio si misura, non si arrotonda a «nessuna»: la prima stesura di questo commento
    # diceva «nessuna dice a quale epic appartiene» mentre il dato — che avevo davanti — ne
    # mostrava una. Il numero qui sotto e' derivato, cosi' la frase non puo' piu' divergere.
    #
    # La menzione di un `RT-FEAT-*` nel testo dell'owner NON basta a dedurre l'appartenenza: E38
    # nomina `RT-FEAT-ACTION-ENGINE` come dipendenza gia' consegnata in v0.1, non come proprio
    # contenuto. Dedurlo avrebbe collegato una feature v0.1 chiusa a un'epic v0.2 aperta.
    orphan = {}
    for entry in data["features"]:
        if entry["release"] in releases and not entry["roadmap"].get("epic"):
            orphan[entry["release"]] = orphan.get(entry["release"], 0) + 1

    no_issue = [e["epic"] for e in epics if e["release"] and not e["issue"]]
    if no_issue:
        lines += [
            "",
            "> ⚠️ **Senza issue dichiarata nell'owner**: " + ", ".join(f"**{e}**" for e in no_issue) +
            ". La epic puo' esistere su GitHub: questo script non parla con la rete, e cio' che "
            "l'owner non dichiara non compare.",
        ]
    if unassigned:
        lines += [
            "",
            "> ⚠️ **Con una sezione ma senza release**: "
            + ", ".join(f"**{e['epic']}**" for e in unassigned) +
            ". La tabella «Le release» non le assegna, e la posizione del testo non e' "
            "un'assegnazione: la decisione manca nell'owner, non qui.",
        ]
    if orphan:
        total = sum(orphan.values())
        bound = sum(1 for e in data["features"]
                    if e["release"] in releases and e["roadmap"].get("epic"))
        detail = " · ".join(f"{rel} **{orphan[rel]}**" for rel in sorted(orphan, key=release_key))
        lines += [
            "",
            f"> ⚠️ **La colonna Feature e' quasi vuota perche' il registry non dichiara l'epic**: "
            f"{total} feature post-v0.1 ({detail}) hanno `roadmap.epic` nullo, "
            f"**{bound}** lo dichiara. Il legame esiste nella prosa di "
            "[`roadmap-post-v0.1.md`](roadmap-post-v0.1.md) e nelle issue, ma non in un campo: "
            "finche' non lo e', questa colonna non puo' dire quali capacita' porta un'epic, e il "
            "conflitto `RT-FEAT-CHARACTER-STATE` / `RT-FEAT-CHAR-TRANSFORMATION` su **E34** non "
            "e' diagnosticabile da nessun controllo.",
        ]
    return wrap_block(marker, lines)


def render_shortlist_features(data):
    marker = SHORTLIST_MARKERS["features"]
    kept = preserved_column(SHORTLIST_FILES["features"], marker, r"`(RT-FEAT-[A-Z0-9-]+)`")
    by_area = {}
    for entry in data["features"]:
        by_area.setdefault(entry["area"], []).append(entry)
    counts = {}
    for entry in data["features"]:
        counts[entry["status"]] = counts.get(entry["status"], 0) + 1
    releases = {}
    for entry in data["features"]:
        releases[entry["release"]] = releases.get(entry["release"], 0) + 1

    lines = [
        f"**{data['count']} feature** · "
        + " · ".join(f"{rel} **{releases[rel]}**" for rel in RELEASE_ORDER if rel in releases)
        + ".",
        "",
        "| Stato | Quante |",
        "|---|--:|",
    ]
    for status in STATUS_ORDER:
        if counts.get(status):
            lines.append(f"| `{status}` | {counts[status]} |")
    for status in sorted(s for s in counts if s in STATUS_OFF_SCALE):
        lines.append(f"| `{status}` (fuori scala) | {counts[status]} |")

    for area in sorted(by_area):
        entries = sorted(by_area[area], key=lambda e: (-STATUS_ORDER.index(e["status"])
                                                       if e["status"] in STATUS_ORDER else 0,
                                                       e["feature_id"]))
        lines += [
            "",
            f"### {area} · {len(entries)}",
            "",
            "| Feature | Rel. | Stato | Gate | Vista | Cosa fissa |",
            "|---|:--:|:--:|--:|:--:|---|",
        ]
        for entry in entries:
            roadmap = entry["roadmap"]
            where = roadmap.get("epic") or roadmap.get("milestone") or (
                "fuori scope" if roadmap.get("out_of_release_scope") else "—")
            lines.append(
                f"| `{entry['feature_id']}` — {entry['title']} | {entry['release']} | "
                f"{entry['status']} | {entry['gates_done']}/{entry['gates_applicable']} | "
                f"{where} | {kept.get(entry['feature_id'], '—')} |"
            )
    return wrap_block(marker, lines)


def render_shortlist_scenarios(data):
    marker = SHORTLIST_MARKERS["scenarios"]
    available = available_capabilities()
    corpus = scenario_corpus()
    blocked, runnable = [], []
    for scenario in corpus:
        missing = [c for c in scenario["requires"] if c not in available]
        (blocked if missing else runnable).append((scenario, missing))

    planned = {}
    for entry in data["features"]:
        for sid in entry["scenarios_planned"]:
            planned.setdefault(sid, entry["feature_id"])

    lines = [
        f"**{len(corpus)} scenari versionati** — misurati su `Scenarios/`: "
        f"**{len(runnable)}** eseguibili · **{len(blocked)}** `BLOCKED` per una capability assente · "
        f"**{len(planned)}** dichiarati `planned` nel registry e non ancora scritti.",
        "",
        "**Capability disponibili oggi**, lette da `RTScenarioSession.cpp` (stanno nel codice, non "
        "nei dati: un JSON che se le dichiarasse da sé produrrebbe il primo verde bugiardo): "
        + " · ".join(f"`{c}`" for c in sorted(available)) + ".",
    ]
    if blocked:
        lines += [
            "",
            "| Scenario `BLOCKED` | Capability che manca |",
            "|---|---|",
        ]
        for scenario, missing in blocked:
            lines.append(
                f"| `{scenario['id']}` | " + " · ".join(f"`{c}`" for c in missing) + " |")
    else:
        lines += ["", "> ✅ **Nessuno scenario del corpus è `BLOCKED`**: ogni `requires` scritto "
                      "trova la sua capability nel codice."]
    if planned:
        lines += [
            "",
            "| Pianificato, non scritto | Feature che lo chiede |",
            "|---|---|",
        ]
        for sid in sorted(planned):
            lines.append(f"| `{sid}` | `{planned[sid]}` |")
    return wrap_block(marker, lines)


def render_shortlist_milestones(data):
    marker = SHORTLIST_MARKERS["milestones"]
    kept = preserved_column(SHORTLIST_FILES["milestones"], marker, r"\*\*(M\d+)\*\*")
    statuses, conflicts = milestone_status()
    by_milestone = {}
    for entry in data["features"]:
        milestone = entry["roadmap"].get("milestone")
        if milestone:
            by_milestone.setdefault(milestone, []).append(entry)

    lines = [
        "| Milestone | Stato | Feature tracciate qui | In una riga |",
        "|---|:--:|--:|---|",
    ]
    for milestone in sorted(set(statuses) | set(by_milestone), key=lambda m: int(m[1:])):
        entries = by_milestone.get(milestone, [])
        tracked = ", ".join(f"`{e['feature_id']}`" for e in sorted(
            entries, key=lambda e: e["feature_id"])) or "—"
        lines.append(
            f"| **{milestone}** | {statuses.get(milestone, '—')} | {tracked} | "
            f"{kept.get(milestone, '—')} |"
        )
    if conflicts:
        lines += [
            "",
            "> ⚠️ **Divergenza nell'owner, non in questa vista.** "
            "[`roadmap-checkpoint.md`](roadmap-checkpoint.md) dichiara la stessa milestone con due "
            "simboli diversi in due tabelle: "
            + " · ".join(f"**{m}** ({a} e {b})" for m, a, b in conflicts)
            + ". Qui è riportata la lettura **più conservativa**; la correzione va fatta là.",
        ]
    return wrap_block(marker, lines)


def render_shortlist_gates(_data):
    """I gate `G1`-`G15`, letti dalla Definition of Done che ne e' l'owner."""
    marker = SHORTLIST_MARKERS["gates"]
    gates = release_gates()
    if not gates:
        return wrap_block(marker, ["> ⚠️ Nessun gate letto da `v0.1-definition-of-done.md` §3."])
    green = sum(1 for _id, _req, state in gates if "✅" in state)
    lines = [
        f"**{len(gates)} gate** · verdi: **{green}**. Stato letto da "
        f"[`v0.1-definition-of-done.md`](v0.1-definition-of-done.md) §3.",
        "",
        "| Gate | Cosa chiede | Stato |",
        "|:--:|---|---|",
    ]
    for gate_id, request, state in gates:
        lines.append(f"| **{gate_id}** | {request} | {state or '—'} |")
    return wrap_block(marker, lines)


def load_editor_sessions():
    """Le sedute in editor. Sorgente dei DATI; `editormap.shortlist.md` ne e' la vista.

    A differenza delle altre quattro shortlist, qui la prosa umana non va **preservata** dal
    markdown: vive nello YAML, che e' la sorgente. Il generatore non ha nulla da restituire
    perche' non ha nulla da distruggere.
    """
    if not os.path.isfile(EDITOR_SESSIONS_YAML):
        return {}
    try:
        import yaml
    except ImportError:
        sys.exit("Serve PyYAML: pip install pyyaml")
    with open(EDITOR_SESSIONS_YAML, encoding="utf-8") as fh:
        return yaml.safe_load(fh) or {}


def pie_entries():
    """Ogni voce del registro PIE con il suo stato, misurato la' dove vive.

    Il marcatore e' il **primo** glifo dell'ultima cella. Cercarlo ovunque nella cella conta
    come verde una voce 🟡 il cui testo cita un ✅ nella propria storia: e' il difetto che il
    comando `awk` del registro ha gia' pagato, ed e' per questo che usa `match` + `substr`.
    """
    entries = {}
    if not os.path.isfile(PIE_REGISTRY):
        return entries
    text = open(PIE_REGISTRY, encoding="utf-8").read()
    for line in text.split("\n"):
        if not line.startswith("| **PIE-"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) < 2:
            continue
        match = re.match(r"\*\*(PIE-[A-Za-z0-9.\-]+)\*\*", cells[0])
        if not match:
            continue
        state = ""
        for char in cells[-1]:
            if char in STATUS_GLYPHS:
                state = char
                break
        entries[match.group(1)] = state
    return entries


def tracked_artifacts(paths):
    """Quali artefatti dichiarati sono davvero tracciati da git.

    L'oracolo e' `git ls-files`, non l'esistenza del file su disco, e la ragione e' concreta:
    `.gitignore` esclude `Content/**/*.uasset` e riammette i singoli file con un `!` esplicito.
    Un asset creato in editor e non aggiunto all'allowlist **non viene tracciato da `git add`,
    in silenzio** — e una seduta che lo desse per committato si dichiarerebbe finita senza esserlo.
    """
    paths = sorted(set(paths))
    if not paths:
        return set()
    try:
        import subprocess
        out = subprocess.run(["git", "ls-files", "--"] + paths, cwd=REPO,
                             capture_output=True, text=True, timeout=30)
        return {line.strip().replace("\\", "/") for line in out.stdout.split("\n") if line.strip()}
    except Exception:
        return set()


def session_state(session, pie, tracked):
    """Stato DERIVATO di una seduta: dalle sue voci PIE e dai suoi artefatti in git.

    Regola dura: un artefatto non tracciato impedisce il verde **qualunque cosa** dicano le voci.
    Una seduta che non dichiara ne' voci ne' artefatti resta senza stato: e' un'informazione
    onesta — il codice sotto non esiste ancora — non un buco da riempire con un simbolo inventato.
    """
    verifies = session.get("verifies") or []
    artifacts = session.get("artifacts") or []
    if not verifies and not artifacts:
        return "—"
    states = [pie.get(v, "") for v in verifies]
    green = sum(1 for s in states if s == "✅")
    partial = sum(1 for s in states if s == "🟡")
    done_artifacts = [a for a in artifacts if a in tracked]
    if green == len(verifies) and len(done_artifacts) == len(artifacts):
        return "✅"
    if green or partial or done_artifacts:
        return "🟡"
    return "⏳"


def checkpoint_status():
    """Stato dei checkpoint, nei DUE spazi di numerazione, tenuti separati.

    `6.3` e' «Input e pianificazione» in **M6** e «Riva» in **E6**: 20 numeri su 22 collidono, e
    per questo la chiave qui e' sempre prefissata. Chi legge passa `M6.3` o `E1.3`, mai `6.3`.

    Un checkpoint dichiarato senza glifo — `| **M8.1** | ... |` — vale ⏳: la riga esiste, il
    lavoro non e' cominciato. Assente del tutto vale invece `None`, che e' un'altra cosa e va
    segnalata a chi lo cita.
    """
    status = {}
    if os.path.isfile(ROADMAP_CHECKPOINT):
        text = open(ROADMAP_CHECKPOINT, encoding="utf-8").read()
        for num, glyph in re.findall(
                r"^\| \*\*M(\d+\.\d+)\*\*\s*([" + STATUS_GLYPHS + r"])?", text, re.M):
            status.setdefault(f"M{num}", glyph or "⏳")
    if os.path.isfile(ROADMAP_V01):
        text = open(ROADMAP_V01, encoding="utf-8").read()
        # Due forme convivono nell'owner: la nuda `| **6.7** |` (94 righe) e la prefissata
        # `| **E21.1** |`, convenzione dal 2026-08-08 (3 righe). Leggerne una sola lascia dei
        # checkpoint senza stato, e una feature che li cita sembra citare il nulla.
        for num, glyph in re.findall(
                r"^\| \*\*E?(\d+\.\d+)\*\*\s*([" + STATUS_GLYPHS + r"])?", text, re.M):
            # In `roadmap-v0.1.md` ogni checkpoint appartiene all'epic che ne apre il numero:
            # `6.7` e' E6.7 per costruzione, e la forma prefissata lo rende dicibile.
            status.setdefault(f"E{num}", glyph or "⏳")
    return status


def resolve_prerequisite(ref, sessions_state, checkpoints, epics, milestones):
    """Un `unblocked_by` e' risolto? Ritorna (risolto, stato_letto, sorgente).

    Le due regole sono diverse, e la differenza e' sostanziale:

    - **checkpoint ed epic**: 🟡 conta come **risolto**. Il prerequisito e' che il *codice* sia
      fatto, e a quel checkpoint manca proprio la verifica che porta la persona in editor.
      Aspettare il ✅ sarebbe aspettare se' stessi.
    - **un'altra seduta**: serve il ✅. Una seduta 🟡 ha l'artefatto a meta', e chi lo estende
      non puo' partire: `U13` estende l'asset che `U1` committa.

    Un riferimento che non risolve ritorna sorgente `None`: e' un difetto da dire, non da
    nascondere dietro un simbolo prudente.
    """
    ref = ref.strip()
    if re.fullmatch(r"U\d+", ref):
        state = sessions_state.get(ref)
        if state is None:
            return False, None, None
        return state == "✅", state, "seduta"
    if re.fullmatch(r"[EM]\d+\.\d+", ref):
        state = checkpoints.get(ref)
        if state is None:
            return False, None, None
        return state in ("✅", "🟡"), state, "checkpoint"
    if re.fullmatch(r"E\d+", ref):
        state = epics.get(ref)
        if state is None:
            return False, None, None
        return state in ("✅", "🟡"), state, "epic"
    if re.fullmatch(r"M\d+", ref):
        state = milestones.get(ref)
        if state is None:
            return False, None, None
        return state in ("✅", "🟡"), state, "milestone"
    return False, None, None


SESSION_FIELDS = {"id", "title", "block", "critical", "produces", "artifacts", "unblocked_by",
                  "shares_setup_with", "verifies", "issues", "done_when", "unblocks",
                  "execution_lane", "steps", "notes"}

# Le due lane del lavoro umano. Non e' una terza sorgente: `editor-sessions.yaml` resta l'owner
# delle sedute, e questo campo dice soltanto *che tipo* di lavoro sono. Il default esiste perche'
# il campo e' nato dopo le sedute: assente = `pie`, che e' cio' che tutte e 21 erano prima.
SESSION_LANES = ("pie", "asset")
SESSION_LANE_DEFAULT = "pie"


def session_lane(session):
    """La lane di una seduta, col default retrocompatibile.

    Una funzione e non `session.get("execution_lane", "pie")` sparso: il default e' una regola —
    «prima del 2026-08-13 ogni seduta era PIE» — e una regola scritta in cinque punti diventa
    cinque regole al primo che ne cambia uno.
    """
    return (session.get("execution_lane") or SESSION_LANE_DEFAULT).strip()


def editor_context():
    """Tutto cio' che serve per ragionare sulle sedute, misurato una volta sola.

    Sono sei letture non banali — il registro PIE, `git ls-files` sugli artefatti (un sottoprocesso),
    e tre file di roadmap riletti a colpi di regex. `validate` e `build_graph` le vogliono entrambi,
    e `build_graph` chiama `validate`: senza questo contesto condiviso lo stesso lavoro si fa due
    volte per ogni `generate`.

    Non e' una cache: e' un parametro. Una cache di modulo avrebbe legato il risultato alla vita del
    processo, e la prima cosa che si rompe e' la verifica di mutazione — che muta il file e
    richiama. Qui chi vuole una misura fresca chiama di nuovo `editor_context()`.
    """
    doc = load_editor_sessions()
    sessions = doc.get("sessions") or []
    pie = pie_entries()
    tracked = tracked_artifacts([a for s in sessions for a in (s.get("artifacts") or [])])
    milestones, milestone_conflicts = milestone_status()
    return {
        "doc": doc,
        "sessions": sessions,
        "pie": pie,
        "tracked": tracked,
        "states": {s.get("id"): session_state(s, pie, tracked) for s in sessions},
        "checkpoints": checkpoint_status(),
        "epics": epic_status(),
        "milestones": milestones,
        "milestone_conflicts": milestone_conflicts,
    }


# Come si stampa una seduta priva di `id`. La stringa e' la stessa che `validate` usa gia' per una
# feature senza id (riga ~272): l'input malformato ha un nome solo in tutto il file. La variante
# `_MD` esiste perche' le viste sono markdown, dove `<senza id>` nudo verrebbe letto come tag HTML e
# sparirebbe dalla pagina — cioe' proprio la riga che serve a vedere il difetto.
NO_SESSION_ID = "<senza id>"
NO_SESSION_ID_MD = f"`{NO_SESSION_ID}`"


def validate_editor_sessions(ctx=None):
    """`editor-sessions.yaml`: la seconda sorgente delle *map, con le stesse pretese della prima.

    Fino al 2026-08-10 questo file non passava di qui. I suoi controlli esistevano — voci `PIE-*`
    fantasma, prerequisiti che non risolvono — ma vivevano dentro il **generatore** e finivano come
    ⚠️ stampati in fondo alla pagina generata. Un avviso dentro la vista che dovrebbe denunciare non
    ferma nessuno e non fa fallire niente: e' la stessa famiglia del dato che nessun consumatore
    legge. Qui diventano exit code.

    La regola meno ovvia e' la dualita' di `unblocks`:

    - `X.unblocked_by` cita la seduta `Y`  ⇒  `Y.unblocks` deve citare `X`, sempre. Chi porta una
      dipendenza deve sapere cosa sblocca: e' la motivazione che la coda stampa per farla adesso.
    - `Y.unblocks` cita la seduta `X`      ⇒  `X.unblocked_by` deve citare `Y`, **salvo** setup
      condiviso **nello stesso blocco**. `U2..U6` (blocco 2) e `U7..U9` (blocco 3) si fanno nella
      stessa apertura di editor: li' la catena e' ordine di lavoro e la coda deve mostrarle
      insieme, non metterne quattro in WAITING.

    Senza l'eccezione questo controllo avrebbe segnalato sei casi legittimi e sarebbe stato
    disattivato al primo passaggio — cioe' non varrebbe piu' nemmeno per il settimo.

    Il vincolo «stesso blocco» non e' decorativo, ed e' costato un falso negativo in review:
    `shares_setup_with` e dipendenza dura **non si escludono**. `U13` (blocco 5) estende l'asset
    che `U1` (blocco 1) committa — l'esempio che l'header del file usa per definire `unblocked_by`
    — e le due condividono anche l'allestimento. Con l'eccezione scritta sul solo setup condiviso,
    togliere `U1` dagli `unblocked_by` di `U13` non produceva **nessun** errore: il controllo
    scusava proprio la dipendenza piu' documentata del file. Due sedute di blocchi diversi non si
    fanno nella stessa apertura, quindi li' la dualita' si pretende sempre.
    """
    # Il contesto arriva da fuori quando chi chiama lo ha gia' misurato (`build_graph`), e viene
    # calcolato qui quando nessuno lo ha fatto. Le sei letture costano un sottoprocesso git e tre
    # riletture di roadmap: farle due volte per `generate` era lavoro doppio senza guadagno.
    # `.get("id")` e non `s["id"]` in `editor_context`: una seduta senza id e' precisamente uno dei
    # difetti che questa funzione esiste per denunciare, e con l'accesso diretto morirebbe prima di
    # arrivare alla riga che lo dice. Un validator che va in traceback sul caso che deve segnalare
    # non lo segnala.
    ctx = ctx or editor_context()
    doc = ctx["doc"]
    sessions = ctx["sessions"]
    if not sessions:
        return [], []

    errors, warnings = [], []
    ids = [s.get("id") for s in sessions]
    id_set = set(ids)
    by_id = {s.get("id"): s for s in sessions}
    blocks = {b.get("id") for b in (doc.get("meta", {}).get("blocks") or [])}

    pie = ctx["pie"]
    states = ctx["states"]
    checkpoints = ctx["checkpoints"]
    epics = ctx["epics"]
    milestones = ctx["milestones"]

    for sid in sorted({i for i in ids if ids.count(i) > 1}):
        errors.append(f"[{sid}] seduta duplicata: l'id non si riusa mai")

    for session in sessions:
        sid = session.get("id") or "<senza id>"
        where = f"[{sid}]"

        if not re.fullmatch(r"U\d+", sid):
            errors.append(f"{where} id non conforme a U<n>")
        if session.get("block") not in blocks:
            errors.append(f"{where} block {session.get('block')!r} assente da meta.blocks: "
                          "la vista lo stamperebbe senza titolo")
        # Il campo e' facoltativo, ma un valore *sbagliato* non e' un'assenza: `execution_lane: PIE`
        # o `editor` passerebbe il default e la seduta finirebbe nella lane opposta senza che nulla
        # lo dica. Si controlla solo quando e' dichiarato.
        if "execution_lane" in session and session_lane(session) not in SESSION_LANES:
            errors.append(f"{where} execution_lane non valida: "
                          f"{session.get('execution_lane')!r} — attese "
                          + " o ".join(f"`{lane}`" for lane in SESSION_LANES))
        for field in session:
            if field not in SESSION_FIELDS:
                errors.append(f"{where} campo non documentato nell'header del file: {field}")

        for field in ("unblocked_by", "unblocks"):
            for ref in session.get(field) or []:
                _resolved, _state, source = resolve_prerequisite(
                    ref, states, checkpoints, epics, milestones)
                if source is None:
                    errors.append(
                        f"{where} {field}: {ref!r} non risolve. Le forme ammesse sono prefissate "
                        "(`M6.3`, `E1.3`, `E8`, `U13`): la forma nuda `CP 6.3` non dice quale dei "
                        "due spazi di numerazione, e 20 numeri su 22 collidono")

        for ref in session.get("shares_setup_with") or []:
            if ref not in id_set:
                errors.append(f"{where} shares_setup_with verso una seduta inesistente: {ref}")
            elif sid not in (by_id[ref].get("shares_setup_with") or []):
                errors.append(f"{where} shares_setup_with cita {ref}, ma {ref} non ricambia: "
                              "lo stesso allestimento non puo' esserlo per una sola delle due")

        for ref in session.get("unblocked_by") or []:
            if ref in id_set and sid not in (by_id[ref].get("unblocks") or []):
                errors.append(f"{where} dichiara di dipendere da {ref}, ma {ref}.unblocks non cita "
                              f"{sid}: la coda perde la ragione per fare {ref} adesso")

        for ref in session.get("unblocks") or []:
            if ref not in id_set:
                continue
            if sid in (by_id[ref].get("unblocked_by") or []):
                continue
            if (ref in (session.get("shares_setup_with") or [])
                    and by_id[ref].get("block") == session.get("block")):
                continue  # stessa apertura, stesso blocco: ordine di lavoro, non un blocco
            shared = ref in (session.get("shares_setup_with") or [])
            perche = ("condividono il setup ma stanno in blocchi diversi, quindi non e' la stessa "
                      "apertura di editor" if shared else "non condividono il setup")
            errors.append(
                f"{where} dichiara di sbloccare {ref}, ma {ref} non lo dichiara fra i suoi "
                f"unblocked_by e le due {perche}. O {ref} dipende davvero da {sid} — e allora lo "
                f"deve dire, o la coda lo mostrera' pronto senza la sua premessa — oppure {sid} "
                f"non lo sblocca")

        for ref in session.get("verifies") or []:
            if ref not in pie:
                errors.append(f"{where} voce PIE inesistente nel registro: {ref} — o l'id e' "
                              "scritto male, o la voce e' stata rinominata: in entrambi i casi "
                              "quella verifica non la esegue nessuno")

    claimed = {}
    for session in sessions:
        for ref in session.get("verifies") or []:
            claimed.setdefault(ref, []).append(session.get("id"))
    for ref, who in sorted(claimed.items()):
        if len(who) > 1:
            # `w or NO_SESSION_ID` e non `w`: l'accesso qui sopra era gia' `.get("id")` dalla PR #398,
            # ma il `None` che ne esce arrivava intatto fino a questo `join` e lo faceva morire in
            # `TypeError`. Correggere l'accesso senza correggere il consumatore aveva solo spostato
            # il traceback di sette righe: `validate` continuava a non arrivare a stampare i tre
            # errori che aveva gia' raccolto sulla stessa seduta.
            warnings.append(f"voce PIE {ref} rivendicata da piu' sedute: "
                            f"{', '.join(w or NO_SESSION_ID for w in who)}")

    return errors, warnings


QUEUE_GROUPS = ["BLOCKING", "READY", "WAITING", "DONE"]


def session_queue(sessions, states, checkpoints, epics, milestones):
    """`My Editor Queue`: i quattro gruppi, DERIVATI dai campi che le sedute gia' dichiarano.

    Nessun campo nuovo: `unblocked_by` dice se si puo' cominciare, `critical` se blocca la v0.1,
    e lo stato dice se e' finita. Il raggruppamento e' una funzione di questi tre — se fosse un
    campo scritto a mano sarebbe la quinta verita' sullo stesso fatto, ed e' esattamente il
    difetto che ha ritirato `roadmap-editor.md` il 2026-08-08.

    Una seduta senza stato derivabile resta fuori dalla coda: il codice sotto non esiste ancora,
    e metterla in WAITING suggerirebbe che basta aspettare.
    """
    queue = {group: [] for group in QUEUE_GROUPS}
    unresolved = []
    for session in sessions:
        # `.get("id")` e non `s["id"]`, come in `editor_context` (vedi il commento sopra, PR #398):
        # `validate` denuncia una seduta senza id, ma `shortlist` e `generate` morivano in `KeyError`
        # sullo stesso input. Cioe' i due comandi che NON sono il gate crashavano proprio sul caso
        # che il gate esiste per segnalare. Una chiave `None` attraversa `states`, `ratios` e
        # `group_of` senza danno; a valle si stampa come NO_SESSION_ID.
        sid = session.get("id")
        state = states.get(sid, "—")
        if state == "—":
            continue
        if state == "✅":
            queue["DONE"].append((session, []))
            continue
        blockers = []
        for ref in session.get("unblocked_by") or []:
            resolved, ref_state, source = resolve_prerequisite(
                ref, states, checkpoints, epics, milestones)
            if source is None:
                unresolved.append((sid, ref))
            if not resolved:
                blockers.append((ref, ref_state))
        if blockers:
            queue["WAITING"].append((session, blockers))
        elif session.get("critical"):
            queue["BLOCKING"].append((session, []))
        else:
            queue["READY"].append((session, []))
    return queue, unresolved


def render_editor_queue(sessions, states, ratios):
    """Le righe della coda. Ogni voce dice cosa sblocca: e' la ragione per farla adesso."""
    checkpoints = checkpoint_status()
    epics = epic_status()
    milestones, _conflicts = milestone_status()
    queue, unresolved = session_queue(sessions, states, checkpoints, epics, milestones)

    counts = " · ".join(f"**{g}** {len(queue[g])}" for g in QUEUE_GROUPS)
    lines = [
        "### My Editor Queue",
        "",
        f"{counts}. **Derivata**, non dichiarata: `unblocked_by` risolto dice se si puo' "
        "cominciare, `critical` se blocca la v0.1, lo stato se e' finita. Un checkpoint 🟡 conta "
        "come risolto — gli manca la verifica che porti tu; una **seduta** prerequisito no, "
        "perche' a meta' non ha ancora prodotto il suo artefatto.",
        "",
    ]
    titles = {
        "BLOCKING": "Blocca la v0.1, e si puo' fare adesso",
        "READY": "Si puo' fare adesso, fuori percorso critico",
        "WAITING": "Aspetta codice",
        "DONE": "Finite",
    }
    for group in QUEUE_GROUPS:
        entries = queue[group]
        lines += [f"**{group}** — *{titles[group]}*", ""]
        if not entries:
            lines += ["- —", ""]
            continue
        for session, blockers in entries:
            sid = session.get("id")
            head = f"- **{sid or NO_SESSION_ID_MD}** · {session.get('title')}"
            if group == "WAITING":
                waits = ", ".join(
                    f"`{ref}`{'' if state is None else ' ' + state}" if state is not None
                    else f"`{ref}` **non risolve**" for ref, state in blockers)
                head += f" — attende {waits}"
            elif group == "DONE":
                head += " ✅"
            else:
                ratio = ratios.get(sid)
                if ratio and ratio != "—":
                    head += f" — {ratio} voci verdi"
                if session.get("unblocks"):
                    head += f" · sblocca {', '.join(session['unblocks'])}"
            lines.append(head)
        lines.append("")
    if unresolved:
        lines += [
            "> ⚠️ **Prerequisiti che non risolvono**: "
            + " · ".join(f"`{sid}` → `{ref}`" for sid, ref in unresolved)
            + ". La forma ammessa e' prefissata (`M6.3`, `E1.3`, `E8`, `U13`): la forma nuda "
              "`CP 6.3` non dice quale dei due spazi di numerazione, e 20 numeri su 22 collidono.",
            "",
        ]
    return lines


def render_shortlist_editor(_data):
    """La vista operativa in editor: sequenza, prerequisiti, artefatti, condizione di chiusura.

    Le quattro cose che nessun altro documento dice. Tutto il resto — esito atteso delle voci,
    stato delle feature, classificazione degli scenari — viene citato per ID e mai ripetuto:
    e' la regola che ha fatto sopravvivere questa vista alla propria prima morte.
    """
    marker = SHORTLIST_MARKERS["editor"]
    doc = load_editor_sessions()
    sessions = doc.get("sessions") or []
    if not sessions:
        return wrap_block(marker, [
            "> ⚠️ Nessuna seduta letta da `editor-sessions.yaml`: la sorgente manca o e' vuota."])

    pie = pie_entries()
    tracked = tracked_artifacts([a for s in sessions for a in (s.get("artifacts") or [])])
    blocks = {b["id"]: b for b in (doc.get("meta", {}).get("blocks") or [])}

    states = {s.get("id"): session_state(s, pie, tracked) for s in sessions}
    counts = {glyph: sum(1 for v in states.values() if v == glyph) for glyph in "✅🟡⏳"}
    unknown = sum(1 for v in states.values() if v == "—")

    summary = (f"**{len(sessions)} sedute** — ✅ **{counts['✅']}** · 🟡 **{counts['🟡']}** · "
               f"⏳ **{counts['⏳']}**")
    if unknown:
        summary += (f" · **{unknown}** senza stato derivabile (non dichiarano ne' voci ne' "
                    f"artefatti: il codice sotto non esiste ancora)")
    ratios = {}
    for session in sessions:
        verifies = session.get("verifies") or []
        green = sum(1 for v in verifies if pie.get(v) == "✅")
        ratios[session.get("id")] = f"{green}/{len(verifies)}" if verifies else "—"

    lines = [
        summary + ".",
        "",
        "Stato **derivato**, mai dichiarato: dalle voci `PIE-*` di "
        "[`../technical/test-manuali-pie.md`](../technical/test-manuali-pie.md) e da `git ls-files` "
        "sugli artefatti. Un artefatto non tracciato impedisce il verde qualunque cosa dicano le voci.",
        "",
    ]
    lines += render_editor_queue(sessions, states, ratios)
    lines += [
        "### Tutte le sedute",
        "",
        "| | Seduta | Lane | Produce | Sbloccata da | Critico | Voci | Stato |",
        "|---|---|:--:|---|---|:--:|:--:|:--:|",
    ]
    for session in sessions:
        verifies = session.get("verifies") or []
        green = sum(1 for v in verifies if pie.get(v) == "✅")
        ratio = f"{green}/{len(verifies)}" if verifies else "—"
        unblocked = ", ".join(session.get("unblocked_by") or []) or "—"
        sid = session.get("id")
        lines.append(
            f"| **{sid or NO_SESSION_ID_MD}** | {session.get('title')} | "
            f"`{session_lane(session).upper()}` | "
            f"{session.get('produces') or '—'} | "
            f"{unblocked} | {'sì' if session.get('critical') else 'no'} | {ratio} | "
            f"{states.get(sid, '—')} |"
        )
    lanes = {}
    for session in sessions:
        lane = session_lane(session)
        lanes[lane] = lanes.get(lane, 0) + 1
    lines += [
        "",
        "**Lane**: " + " · ".join(f"`{lane.upper()}` **{lanes[lane]}**"
                                 for lane in SESSION_LANES if lane in lanes) +
        ". `ASSET` significa che l'uscita e' un asset da costruire e committare, `PIE` che e' un "
        "verdetto da dare guardando il gioco. Non e' l'evidenza: U7 e' `ASSET` **e** verifica due "
        "voci `PIE-*`. Serve a rispondere a una domanda sola — *cosa mi serve per farla, il gioco "
        "che gira o gli asset che non ho ancora?*",
    ]

    seen_blocks = []
    for session in sessions:
        block_id = session.get("block")
        if block_id not in seen_blocks:
            seen_blocks.append(block_id)
            block = blocks.get(block_id, {})
            lines += ["", f"### Blocco {block_id} — {block.get('title', '—')}"]
            if block.get("note"):
                lines += ["", f"*{block['note']}*"]
        sid = session.get("id")
        lines += ["", f"#### {sid or NO_SESSION_ID_MD} · {session.get('title')} "
                      f"{states.get(sid, '—')}", ""]

        head = [f"**Sbloccata da**: {', '.join(session.get('unblocked_by') or []) or '—'}"]
        if session.get("shares_setup_with"):
            head.append("**Preparazione condivisa con**: "
                        + ", ".join(session["shares_setup_with"]))
        head.append(f"**Percorso critico**: {'sì' if session.get('critical') else 'no'}")
        lines.append(" · ".join(head))
        lines.append(f"**Produce**: {session.get('produces') or '—'}")

        artifacts = session.get("artifacts") or []
        if artifacts:
            lines.append("**Artefatti**: " + " · ".join(
                f"`{a}` {'✅' if a in tracked else '⏳'}" for a in artifacts))
        verifies = session.get("verifies") or []
        if verifies:
            lines.append("**Verifichi**: " + " · ".join(
                f"`{v}` {pie.get(v) or '❓'}" for v in verifies))
        if session.get("done_when"):
            lines.append(f"**Finita quando**: {session['done_when']}")
        if session.get("unblocks"):
            lines.append(f"**Sblocca**: {', '.join(session['unblocks'])}")
        if session.get("steps"):
            lines += ["", session["steps"].rstrip()]
        if session.get("notes"):
            lines += ["", f"> {session['notes'].strip()}"]

    declared = {v for s in sessions for v in (s.get("verifies") or [])}
    ghosts = sorted(v for v in declared if v not in pie)
    if ghosts:
        lines += [
            "",
            "> ⚠️ **Voci dichiarate da una seduta e assenti dal registro**: "
            + " · ".join(f"`{g}`" for g in ghosts)
            + ". O l'ID e' scritto male in `editor-sessions.yaml`, o la voce e' stata rinominata "
              "nel registro: in entrambi i casi quella verifica oggi non viene eseguita da nessuno.",
        ]
    orphans = sorted(v for v in pie if v not in declared)
    if orphans:
        groups = {}
        for name in orphans:
            # Famiglia, non ID: `PIE-BU2b` e `PIE-BU3c` sono la stessa cosa contata due volte.
            family = re.match(r"PIE-([A-Za-z]+)", name)
            prefix = f"PIE-{family.group(1)}" if family else name
            groups[prefix] = groups.get(prefix, 0) + 1
        lines += [
            "",
            f"> **{len(orphans)} voci del registro non stanno in nessuna seduta** — "
            + " · ".join(f"`{p}-*` {n}" for p, n in sorted(groups.items()))
            + ". Non e' per forza un difetto (le `PIE-VIS-*` hanno il proprio scenario, le "
              "`PIE-STATE-*` verificano un sistema che non esiste), ma una voce che non sta in una "
              "seduta non viene eseguita mai: e' la ragione per cui questo conteggio e' qui.",
        ]
    return wrap_block(marker, lines)


EXEC_HARD = {"requires", "requires_capability"}
EXEC_SOFT = {"follows", "related"}
EXEC_TRACE = {"provides", "implements", "verifies"}
EXEC_RELATIONS = EXEC_HARD | EXEC_SOFT | EXEC_TRACE
# Nomi che il generatore DERIVA e che quindi non si scrivono nel source. Vivono in un dizionario
# proprio, e non in un `if` dentro il loop, perche' il controllo va fatto prima di «tipo
# sconosciuto»: sono per definizione fuori da `EXEC_RELATIONS`.
EXEC_DERIVED = {
    "blocks": "e' l'inversa di `requires`, e due direzioni scritte a mano divergono al primo edit",
    "converges_into": "un junction si riconosce dal numero di ingressi hard, non si dichiara",
}
EXEC_SCHEMA_VERSION = 1


def load_execution_graph():
    """La topologia del lavoro. Assente = non e' un errore: il grafo e' una fetta opzionale."""
    if not os.path.isfile(EXECUTION_GRAPH):
        return {}
    try:
        import yaml
    except ImportError:
        sys.exit("Serve PyYAML: pip install pyyaml")
    with open(EXECUTION_GRAPH, encoding="utf-8") as fh:
        return yaml.safe_load(fh) or {}


def checkpoint_declared():
    """I checkpoint per cui l'owner DICHIARA un glifo, distinti da quelli che lo ricevono a default.

    `checkpoint_status()` restituisce ⏳ anche quando la riga non porta simbolo, e per il suo uso
    — i prerequisiti delle sedute — e' la scelta giusta: la convenzione di `roadmap-v0.1.md` e'
    che il glifo si scrive alla chiusura, quindi l'assenza vale «non chiuso».

    Per l'execution graph quella fusione non regge, e il numero dice perche': su **100**
    checkpoint di epic l'owner dichiara il glifo per **18**. Degli 82 restanti, **65** hanno una
    issue GitHub chiusa che nessuno ha annotato. Leggere quel ⏳ come una misura significherebbe
    dipingere di BLOCKED meta' della roadmap sulla base di un default — e un BLOCKED falso in una
    vista che esiste per dire cosa fare adesso e' peggio di un UNKNOWN onesto.

    Questa funzione non cambia il comportamento esistente: chi vuole il default continua a
    chiamare `checkpoint_status()`. Serve a sapere **se ci si puo' fidare** del valore.
    """
    declared = set()
    for path, pattern in ((ROADMAP_CHECKPOINT, r"^\| \*\*M(\d+\.\d+)\*\*\s*([" + STATUS_GLYPHS + r"])"),
                          (ROADMAP_V01, r"^\| \*\*E?(\d+\.\d+)\*\*\s*([" + STATUS_GLYPHS + r"])")):
        if not os.path.isfile(path):
            continue
        prefix = "M" if path == ROADMAP_CHECKPOINT else "E"
        text = open(path, encoding="utf-8").read()
        for num, _glyph in re.findall(pattern, text, re.M):
            declared.add(f"{prefix}{num}")
    return declared


def execution_node_state(node, ctx, declared):
    """Stato di un nodo e **da dove viene**, o `None` quando nessuna fonte offline lo sa.

    Tre casi, in ordine di affidabilita':
      - `session:U<n>`  -> `session_state()`, che misura voci PIE e `git ls-files`. Affidabile.
      - `issue:` con `checkpoint:` **dichiarato** -> il glifo dell'owner. Affidabile.
      - tutto il resto   -> `None`. Una issue standalone non ha sorgente offline, e il generatore
        non parla con la rete: fingere `⏳` la farebbe sembrare non cominciata anche quando e'
        chiusa da settimane.
    """
    nid = node["id"]
    if nid.startswith("session:"):
        sid = nid.split(":", 1)[1]
        state = ctx["states"].get(sid)
        return (state, f"session:{sid}") if state else (None, None)
    cp = node.get("checkpoint")
    if cp and cp in declared:
        return ctx["checkpoints"].get(cp), f"checkpoint:{cp}"
    return None, None


def build_execution(ctx, registry, corpus=None, available=None):
    """Nodi, archi, capability e conteggi dell'execution graph. Tutto derivato, niente dichiarato.

    ## Readiness: si calcola sui PREREQUISITI, non sul lavoro proprio

    `READY` qui significa **«nessun prerequisito hard lo trattiene»**, non «da fare». La
    distinzione conta perche' lo stato proprio di un nodo spesso non e' misurabile offline, e
    aspettarlo avrebbe reso `UNKNOWN` anche i nodi che nulla blocca — cioe' avrebbe tolto la
    risposta alla sola domanda per cui questa vista esiste.

    Ordine di decisione, dal piu' forte al piu' debole:
      1. il nodo e' ✅            -> `DONE`
      2. un prerequisito hard e' NOTO non risolto:
           - e' una seduta `asset` -> `WAITING_FOR_ASSET`
           - e' una seduta `pie`   -> `WAITING_FOR_PIE`
           - altrimenti            -> `BLOCKED`
      3. un prerequisito hard ha stato ignoto -> `UNKNOWN` (non `BLOCKED`: non saperlo non e'
         un blocco, ed e' il punto in cui una dashboard comincia a mentire)
      4. altrimenti -> `READY`

    `follows` e `related` non entrano in nessuno dei quattro rami. Due test lo pinnano.
    """
    doc = load_execution_graph()
    if not doc:
        return None

    declared = checkpoint_declared()
    lanes = [dict(l) for l in (doc.get("execution_lanes") or [])]
    groups = [dict(g) for g in (doc.get("domain_groups") or [])]
    nodes = doc.get("nodes") or []
    relations = doc.get("relations") or []

    # Feature per issue: l'inversa delle `issues:` del registry. Il mapping esiste gia' la', e
    # riscriverlo nel source dell'execution graph avrebbe creato la copia che diverge.
    feature_of_issue = {}
    for entry in registry["features"]:
        for num in entry.get("issues") or []:
            feature_of_issue.setdefault(num, []).append(entry["feature_id"])

    lane_of_session = {s.get("id"): session_lane(s) for s in ctx["sessions"]}
    by_id, out = {}, []
    for node in nodes:
        nid = node["id"]
        kind = nid.split(":", 1)[0]
        ref = nid.split(":", 1)[1]
        if kind == "issue":
            lane = node.get("execution_lane") or "code"
        elif kind == "session":
            lane = lane_of_session.get(ref, SESSION_LANE_DEFAULT)
        else:
            lane = node.get("execution_lane")
        state, source = execution_node_state(node, ctx, declared)
        record = {
            "id": nid, "kind": kind, "ref": ref,
            "release": node.get("release"),
            "execution_lane": lane,
            "domain_group": node.get("domain_group"),
            "checkpoint": node.get("checkpoint"),
            "state": state,
            "state_source": source,
            "state_note": None if source else (
                "nessuna fonte offline: il checkpoint non dichiara il glifo"
                if node.get("checkpoint") else "issue senza checkpoint: lo stato vive su GitHub"),
            "feature_ids": sorted(feature_of_issue.get(int(ref), [])) if kind == "issue" and ref.isdigit() else [],
            "provides": list(node.get("provides") or []),
            "incoming": [], "outgoing": [],
            "readiness": None, "readiness_reason": None,
        }
        by_id[nid] = record
        out.append(record)

    edges = []
    for rel in relations:
        edge = {
            "from": rel.get("from"), "to": rel.get("to"), "type": rel.get("type"),
            "hard": rel.get("type") in EXEC_HARD,
            "rationale": (rel.get("rationale") or "").strip() or None,
        }
        edges.append(edge)
        if edge["from"] in by_id:
            by_id[edge["from"]]["outgoing"].append(edge)
        if edge["to"] in by_id:
            by_id[edge["to"]]["incoming"].append(edge)

    # Capability: i consumatori veri sono gli scenari del corpus, che le dichiarano in
    # `requires`. Il provider arriva da qui — ed e' l'unica cosa che il repository non sapeva.
    # Passati da `build_graph`, che li usa comunque per le proprie chiavi: senza, ogni
    # `generate` faceva due `os.walk` di `Scenarios/` con ~73 `json.load` e due letture del
    # sorgente delle capability. E' lo stesso motivo per cui `editor_ctx` viaggia come parametro.
    available = available if available is not None else available_capabilities()
    scenario_consumers = {}
    for scenario in (corpus if corpus is not None else scenario_corpus()):
        for cap in scenario.get("requires") or []:
            scenario_consumers.setdefault(cap, []).append(scenario["id"])
    capabilities = {}
    for edge in edges:
        for side in ("from", "to"):
            if str(edge[side]).startswith("capability:"):
                capabilities.setdefault(edge[side].split(":", 1)[1],
                                        {"providers": [], "consumers": []})
    for edge in edges:
        if edge["type"] == "provides" and str(edge["to"]).startswith("capability:"):
            capabilities[edge["to"].split(":", 1)[1]]["providers"].append(edge["from"])
        if edge["type"] == "requires_capability" and str(edge["from"]).startswith("capability:"):
            capabilities[edge["from"].split(":", 1)[1]]["consumers"].append(edge["to"])

    def resolved(nid):
        """`True` risolto · `False` noto non risolto · `None` non misurabile offline.

        I glifi coperti sono tutti e quattro quelli di `STATUS_GLYPHS`: `⌫` («rimosso dal repo»)
        vale **non risolto**, come gia' fa `resolve_prerequisite()` per le sedute, che accetta
        solo `✅` e `🟡`. Senza, un glifo *dichiarato* dall'owner ricadeva nel `return None`
        finale e il nodo usciva «stato non misurabile offline» avendo in JSON `state: "⌫"` e uno
        `state_source` valorizzato — l'artefatto che contraddice se stesso. Nessun checkpoint usa
        `⌫` oggi: e' una latenza, e costa una riga chiuderla.
        """
        record = by_id.get(nid)
        if record is None:
            return None
        if record["state"] == "✅":
            return True
        if record["state"] in ("🟡", "⏳", "⌫"):
            return False
        return None

    for record in out:
        if record["state"] == "✅":
            record["readiness"], record["readiness_reason"] = "DONE", record["state_source"]
            continue
        blocking, unknown = [], []
        for edge in record["incoming"]:
            if not edge["hard"]:
                continue
            origin = edge["from"]
            if str(origin).startswith("capability:"):
                name = origin.split(":", 1)[1]
                cap = capabilities.get(name, {})
                # Disponibile per due strade indipendenti: il runtime la elenca gia'
                # (`RTScenarioSession.cpp`), oppure un provider del grafo e' chiuso. La prima e'
                # la verita' del codice e vince: una capability che il gioco possiede non torna
                # indisponibile perche' il grafo non ne conosce il provider.
                if name not in available and not any(
                        resolved(p) is True for p in cap.get("providers") or []):
                    blocking.append(origin)
                continue
            verdict = resolved(origin)
            if verdict is False:
                blocking.append(origin)
            elif verdict is None:
                unknown.append(origin)
        if blocking:
            # `WAITING_FOR_*` dice **chi deve intervenire**, quindi lo si afferma solo quando
            # l'attesa e' tutta di quella persona: se anche un solo prerequisito e' codice, il
            # nodo e' `BLOCKED` e mandare qualcuno in editor non lo sbloccherebbe.
            #
            # Prima decideva `blocking[0]`, cioe' l'ordine in cui le relazioni comparivano nello
            # YAML: riordinare due stanze indipendenti cambiava il verdetto della dashboard e il
            # conteggio `by_readiness` che un test pinna. Un verdetto non puo' dipendere da come
            # e' impaginato il source.
            # `blocking_lanes`, non `lanes`: quel nome appartiene gia' alla tassonomia serializzata
            # poche righe piu' su, e riusarlo qui la sostituiva con un set — che `json.dump` non
            # sa scrivere. Il gate l'ha preso al primo `generate`.
            blocking_lanes = {by_id.get(b, {}).get("execution_lane") for b in blocking}
            if blocking_lanes == {"asset"}:
                record["readiness"] = "WAITING_FOR_ASSET"
            elif blocking_lanes <= {"asset", "pie"}:
                record["readiness"] = "WAITING_FOR_PIE"
            else:
                record["readiness"] = "BLOCKED"
            record["readiness_reason"] = "aspetta " + ", ".join(sorted(blocking))
        elif unknown:
            record["readiness"] = "UNKNOWN"
            record["readiness_reason"] = ("stato non misurabile offline di " + ", ".join(unknown))
        else:
            record["readiness"] = "READY"
            record["readiness_reason"] = ("nessun prerequisito hard"
                                          if not any(e["hard"] for e in record["incoming"])
                                          else "tutti i prerequisiti hard sono risolti")

    cap_out = []
    for name, data in sorted(capabilities.items()):
        cap_out.append({
            "id": name,
            "providers": data["providers"],
            "consumers": data["consumers"],
            "scenario_consumers": sorted(scenario_consumers.get(name, [])),
            "available": name in available or any(resolved(p) is True for p in data["providers"]),
            "available_source": "runtime" if name in available else "provider",
        })

    def tally(key):
        counts = {}
        for record in out:
            counts[record[key] or "—"] = counts.get(record[key] or "—", 0) + 1
        return dict(sorted(counts.items()))

    return {
        "schema_version": doc.get("schema_version"),
        "status": (doc.get("meta") or {}).get("status"),
        "lanes": lanes,
        "domain_groups": groups,
        "nodes": out,
        "edges": edges,
        "capabilities": cap_out,
        "stats": {
            "nodes": len(out),
            "edges": len(edges),
            "hard_edges": sum(1 for e in edges if e["hard"]),
            "soft_edges": sum(1 for e in edges if e["type"] in EXEC_SOFT),
            "by_readiness": tally("readiness"),
            "by_execution_lane": tally("execution_lane"),
            "by_domain_group": tally("domain_group"),
            "by_release": tally("release"),
            "state_unknown": sum(1 for r in out if r["state"] is None),
        },
    }


def validate_execution_graph(ctx=None, registry=None):
    """`execution-graph.yaml`: riferimenti risolvibili, cicli hard, tassonomia esaustiva.

    Le due regole meno ovvie, e sono quelle che il file esiste per far rispettare:

    - **`blocks` non si autorializza.** E' l'inversa di `requires` e il generatore la deriva.
      Scriverle entrambe produce due verita' che divergono al primo edit di una sola, ed e' il
      difetto piu' facile da introdurre in un grafo scritto a mano.
    - **`follows` fra due sedute e' vietato.** `editor-sessions.yaml` distingue gia' la
      dipendenza dura (`unblocked_by`) dall'ordine di lavoro nello stesso allestimento
      (`unblocks` senza reciproco, stesso blocco) e ha un validator che lo verifica. Riscrivere
      quell'ordine qui lo dichiarerebbe due volte, in due file che non si controllano a vicenda.

    Verifica anche che i `domain_groups` coprano **tutte** le `area` del registry e nessuna due
    volte: senza, un'area nuova non finirebbe in nessun gruppo e i suoi nodi sparirebbero dai
    filtri senza che niente lo dica.
    """
    doc = load_execution_graph()
    if not doc:
        return [], []
    ctx = ctx or editor_context()
    registry = registry or load_registry()
    errors, warnings = [], []

    if doc.get("schema_version") != EXEC_SCHEMA_VERSION:
        errors.append(f"[execution-graph] schema_version {doc.get('schema_version')!r} "
                      f"sconosciuta: attesa {EXEC_SCHEMA_VERSION}")

    lanes = {l.get("id") for l in (doc.get("execution_lanes") or [])}
    groups = {g.get("id") for g in (doc.get("domain_groups") or [])}

    mapped = {}
    for group in doc.get("domain_groups") or []:
        for area in group.get("feature_areas") or []:
            if area in mapped:
                errors.append(f"[execution-graph] area {area!r} mappata due volte: "
                              f"{mapped[area]} e {group.get('id')}")
            mapped[area] = group.get("id")
    areas = {entry["area"] for entry in registry["features"]}
    for area in sorted(areas - set(mapped)):
        errors.append(f"[execution-graph] area {area!r} del registry non appartiene a nessun "
                      "domain_group: i suoi nodi sparirebbero dai filtri")
    for area in sorted(set(mapped) - areas):
        warnings.append(f"[execution-graph] domain_group mappa l'area {area!r}, che il registry "
                        "non usa: e' un mapping senza soggetto")

    session_ids = {s.get("id") for s in ctx["sessions"]}
    checkpoints = ctx["checkpoints"]
    seen, node_ids = set(), set()
    for node in doc.get("nodes") or []:
        nid = node.get("id")
        where = f"[execution-graph {nid}]"
        if nid in seen:
            errors.append(f"{where} nodo duplicato")
        seen.add(nid)
        node_ids.add(nid)
        if not re.fullmatch(r"(issue:\d+|session:U\d+|epic:E\d+|capability:[A-Za-z0-9_]+)", nid or ""):
            errors.append(f"{where} id non conforme: attesi issue:<n>, session:U<n>, "
                          "epic:E<n>, capability:<Nome>")
            continue
        if node.get("execution_lane") and node["execution_lane"] not in lanes:
            errors.append(f"{where} execution_lane {node['execution_lane']!r} sconosciuta")
        if node.get("domain_group") and node["domain_group"] not in groups:
            errors.append(f"{where} domain_group {node['domain_group']!r} sconosciuto")
        if node.get("release") and node["release"] not in RELEASES:
            errors.append(f"{where} release {node['release']!r} non valida")
        if nid.startswith("session:") and nid.split(":", 1)[1] not in session_ids:
            errors.append(f"{where} seduta inesistente in editor-sessions.yaml")
        cp = node.get("checkpoint")
        if cp and cp not in checkpoints:
            errors.append(f"{where} checkpoint {cp!r} che nessun owner dichiara")
        for field in ("state", "status", "readiness", "open", "closed"):
            if field in node:
                errors.append(f"{where} campo {field!r}: lo stato non e' topologia e non si "
                              "scrive qui — lo deriva il generatore dagli owner")

    adjacency = {}
    for rel in doc.get("relations") or []:
        origin, target, kind = rel.get("from"), rel.get("to"), rel.get("type")
        where = f"[execution-graph {origin} -> {target}]"
        # I nomi derivati si controllano PRIMA di «tipo sconosciuto», e non dopo: non stanno in
        # `EXEC_RELATIONS`, quindi il controllo generico scattava per primo e il messaggio
        # specifico non e' mai stato raggiungibile. La regola sopravviveva per caso — e sarebbe
        # sparita del tutto il giorno in cui qualcuno avesse aggiunto `blocks` all'insieme.
        if kind in EXEC_DERIVED:
            errors.append(f"{where} {kind!r} e' derivata e non si autorializza: "
                          + EXEC_DERIVED[kind])
            continue
        if kind not in EXEC_RELATIONS:
            errors.append(f"{where} tipo di relazione sconosciuto: {kind!r}")
            continue
        for side, ref in (("from", origin), ("to", target)):
            if str(ref).startswith("capability:"):
                continue
            if ref not in node_ids:
                errors.append(f"{where} il lato {side} non e' un nodo dichiarato: {ref!r}")
        if origin == target:
            errors.append(f"{where} un nodo non puo' richiedere se stesso")
        if kind == "follows" and str(origin).startswith("session:") and str(target).startswith("session:"):
            errors.append(
                f"{where} `follows` fra due sedute e' gia' dichiarato da editor-sessions.yaml "
                "(`unblocks` senza reciproco, stesso blocco): qui sarebbe la seconda verita'")
        if kind == "follows" and not (rel.get("rationale") or "").strip():
            warnings.append(f"{where} `follows` senza `rationale`: un ordine consigliato che non "
                            "dice perche' diventa indistinguibile da un blocco dimenticato")
        if kind in EXEC_HARD:
            adjacency.setdefault(origin, []).append(target)

    colour = {}

    def visit(node, trail):
        if colour.get(node) == 2:
            return
        if colour.get(node) == 1:
            errors.append("[execution-graph] ciclo fra dipendenze hard: "
                          + " -> ".join(trail + [node]))
            return
        colour[node] = 1
        for nxt in adjacency.get(node, []):
            visit(nxt, trail + [node])
        colour[node] = 2

    for node in list(adjacency):
        visit(node, [])

    # Il collo di bottiglia della vista, misurato invece che intuito. Un nodo il cui checkpoint
    # non dichiara il glifo non ha stato offline, e la sua ignoranza si propaga a valle: ogni
    # discendente diventa UNKNOWN. Il warning nomina i checkpoint da annotare, perche' la
    # correzione e' una riga nell'owner e non un cambio di modello.
    declared = checkpoint_declared()
    silent = sorted({node["checkpoint"] for node in (doc.get("nodes") or [])
                     if node.get("checkpoint") and node["checkpoint"] not in declared})
    if silent:
        warnings.append(
            f"[execution-graph] {len(silent)} checkpoint citati dal grafo non dichiarano il "
            f"glifo in roadmap-v0.1.md ({', '.join(silent)}): i loro nodi restano UNKNOWN e "
            "l'ignoranza si propaga a valle. Su 100 checkpoint di epic l'owner ne annota 18, e "
            "65 dei restanti hanno gia' una issue chiusa")

    providers = {r.get("to") for r in (doc.get("relations") or []) if r.get("type") == "provides"}
    # Misurata una volta: dentro il loop questa chiamata riapriva e ri-regexava
    # `RTScenarioSession.cpp` una volta per ogni relazione `requires_capability`.
    available = available_capabilities()
    for rel in doc.get("relations") or []:
        if rel.get("type") == "requires_capability" and rel.get("from") not in providers:
            name = str(rel.get("from")).split(":", 1)[1]
            if name not in available:
                warnings.append(f"[execution-graph] la capability {name!r} non e' disponibile a "
                                "runtime e nessun nodo dichiara di fornirla: chi la aspetta "
                                "resta bloccato senza che il grafo sappia dire da chi")
    return errors, warnings


def build_graph(registry):
    """Il contratto delle viste non-feature: diagnostica, roadmap, sedute, PIE, scenari.

    Ogni valore qui dentro e' gia' calcolato da una funzione che esiste e che una shortlist usa:
    questa non e' una seconda derivazione, e' un secondo **consumatore**. La regola che lo tiene
    onesto e' che nessun campo nasce qui — se un dato non ha gia' una sorgente misurabile, non
    entra.

    Fuori di proposito: il commit corrente. Metterlo dentro farebbe fallire `--check` dopo ogni
    commit, trasformando un gate utile in rumore da ignorare. La provenienza la tiene git.
    """
    # Una misura sola, condivisa con `validate`: le sei letture sotto includono un sottoprocesso
    # `git ls-files` e tre riletture di roadmap, e prima venivano fatte due volte per ogni
    # `generate` — una qui e una dentro `validate`.
    ctx = editor_context()
    errors, warnings = validate(registry, editor_ctx=ctx)
    # Le due misure che servono sia alle chiavi di questo dizionario sia all'execution graph:
    # `scenario_corpus()` cammina `Scenarios/` e apre ~73 JSON, `available_capabilities()` legge
    # il sorgente delle capability. Una volta sola, e poi passate.
    corpus = scenario_corpus()
    available = available_capabilities()
    pie = ctx["pie"]
    sessions = ctx["sessions"]
    tracked = ctx["tracked"]
    states = ctx["states"]
    checkpoints = ctx["checkpoints"]
    epics = ctx["epics"]
    milestones, milestone_conflicts = ctx["milestones"], ctx["milestone_conflicts"]
    queue, unresolved = session_queue(sessions, states, checkpoints, epics, milestones)
    group_of = {s.get("id"): group for group in QUEUE_GROUPS for s, _b in queue[group]}

    graph_sessions = []
    for session in sessions:
        sid = session.get("id")
        prerequisites = []
        for ref in session.get("unblocked_by") or []:
            resolved, state, source = resolve_prerequisite(
                ref, states, checkpoints, epics, milestones)
            prerequisites.append({"ref": ref, "resolved": resolved,
                                  "state": state, "kind": source})
        artifacts = session.get("artifacts") or []
        graph_sessions.append({
            "id": sid,
            "title": session.get("title"),
            "block": session.get("block"),
            "critical": bool(session.get("critical")),
            "execution_lane": session_lane(session),
            "state": states.get(sid, "—"),
            "queue_group": group_of.get(sid),
            "produces": (session.get("produces") or "").strip(),
            "artifacts": [{"path": a, "tracked": a in tracked} for a in artifacts],
            "unblocked_by": prerequisites,
            "unblocks": session.get("unblocks") or [],
            "shares_setup_with": session.get("shares_setup_with") or [],
            "verifies": [{"id": v, "state": pie.get(v)} for v in (session.get("verifies") or [])],
            "issues": session.get("issues") or [],
            "done_when": (session.get("done_when") or "").strip(),
        })

    return {
        "_generated": "NON EDITARE: generato da scripts/feature_registry.py generate, "
                      "misurando le sorgenti citate in `sources`",
        "sources": {
            "features": "docs/roadmap/feature-registry.yaml",
            "editor_sessions": "docs/roadmap/editor-sessions.yaml",
            "pie_registry": "docs/technical/test-manuali-pie.md",
            "release_gates": "docs/roadmap/v0.1-definition-of-done.md",
            "epics": "docs/roadmap/roadmap-v0.1.md",
            "milestones": "docs/roadmap/roadmap-checkpoint.md",
            "scenarios": "Scenarios/",
            "execution_graph": "docs/roadmap/execution-graph.yaml",
            "post_v01_releases": "docs/roadmap/roadmap-post-v0.1.md",
        },
        "meta": jsonable(registry.get("meta") or {}),
        "project": jsonable((registry.get("meta") or {}).get("project") or {}),
        "diagnostics": {
            "errors": errors,
            "warnings": warnings,
            "unresolved_prerequisites": [{"session": sid, "ref": ref} for sid, ref in unresolved],
            "milestone_conflicts": [
                {"milestone": m, "declared": a, "also": b} for m, a, b in milestone_conflicts],
        },
        "release_gates": [{"id": gid, "request": req, "state": state}
                          for gid, req, state in release_gates()],
        "epics": epics,
        "milestones": milestones,
        "checkpoints": checkpoints,
        "editor_sessions": graph_sessions,
        "editor_queue": {group: [s.get("id") for s, _b in queue[group]] for group in QUEUE_GROUPS},
        "pie_entries": [{"id": k, "state": v} for k, v in sorted(pie.items())],
        "scenarios": corpus,
        "capabilities": sorted(available),
        "execution": build_execution(ctx, registry, corpus=corpus, available=available),
    }


def apply_shortlist(data, check=False):
    """Riscrive i blocchi marcati nelle cinque shortlist. Fuori dai marcatori non tocca nulla."""
    renderers = {
        "epics": render_shortlist_epics,
        "releases": render_shortlist_releases,
        "features": render_shortlist_features,
        "scenarios": render_shortlist_scenarios,
        "milestones": render_shortlist_milestones,
        "gates": render_shortlist_gates,
        "editor": render_shortlist_editor,
    }
    changed, missing = [], []
    for key, path in SHORTLIST_FILES.items():
        marker = SHORTLIST_MARKERS[key]
        if not os.path.isfile(path):
            missing.append(f"{os.path.relpath(path, REPO)} non esiste")
            continue
        original = open(path, encoding="utf-8").read()
        begin, end = f"<!-- {marker}:BEGIN -->", f"<!-- {marker}:END -->"
        if begin not in original or end not in original:
            missing.append(
                f"{os.path.relpath(path, REPO)} non ha il blocco {marker}: nulla da generare")
            continue
        block = renderers[key](data)
        pattern = re.compile(re.escape(begin) + r".*?" + re.escape(end), re.S)
        text = pattern.sub(lambda _m: block, original)
        if text != original:
            # Un file puo' ospitare piu' blocchi (milestone e gate): si riporta una volta sola.
            ref = os.path.relpath(path, REPO).replace("\\", "/")
            if ref not in changed:
                changed.append(ref)
            if not check:
                with open(path, "w", encoding="utf-8", newline="\n") as fh:
                    fh.write(text)
    return changed, missing


# ---------------------------------------------------------------------------

def write_if_needed(path, content, check):
    current = open(path, encoding="utf-8").read() if os.path.isfile(path) else None
    if current == content:
        return False
    if not check:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8", newline="\n") as fh:
            fh.write(content)
    return True


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("command",
                        choices=["validate", "generate", "wiki", "workbook", "suite",
                                 "shortlist", "deploy", "report"])
    parser.add_argument("--wiki-root", help="radice del clone della Wiki (deploy flat)")
    parser.add_argument("--check", action="store_true", help="non scrivere: fallisci se disallineato")
    parser.add_argument("--write", action="store_true",
                        help="`deploy`: scrive davvero nel clone della Wiki (default: sola lettura)")
    parser.add_argument("--run-log", metavar="PATH",
                        help="`suite`: log di una run dell'automation. Confronta ESEGUITI e DICHIARATI "
                             "nello stesso perimetro ed esce 1 se qualcuno manca (#486)")
    args = parser.parse_args()

    registry = load_registry()
    data = build_json(registry)

    if args.command == "validate":
        errors, warnings = validate(registry, args.wiki_root)
        print(f"Feature: {data['count']}")
        print(f"  riferimenti wiki   : {sum(len(f['wiki_refs']) for f in data['features'])}")
        print(f"  riferimenti roadmap: "
              f"{sum(1 for f in data['features'] if f['roadmap']['epic'] or f['roadmap']['milestone'])}")
        print(f"  riferimenti scenari: {sum(len(f['scenarios']) for f in data['features'])} "
              f"(+{sum(len(f['scenarios_planned']) for f in data['features'])} pianificati)")
        print(f"  riferimenti test   : {sum(len(f['tests']) for f in data['features'])}")
        for warning in warnings:
            print(f"WARN  {warning}")
        for error in errors:
            print(f"ERROR {error}")
        # Fuori dal conteggio dei warning, perche' non e' una lacuna del progetto: dice quanto ha
        # potuto guardare *questa* esecuzione. Senza, un validate senza `--wiki-root` sarebbe verde
        # su riferimenti che non ha aperto — da D-076 sono tutte le pagine di gioco.
        if not args.wiki_root:
            unchecked = {r for f in registry.get("features") or []
                         for r in (f.get("wiki_refs") or []) if r.startswith("wiki:")}
            if unchecked:
                print(f"NOTA  {len(unchecked)} riferimenti `wiki:` non verificati: passa "
                      "--wiki-root <clone> perche' la loro esistenza venga controllata")
        print(f"\nerrori: {len(errors)} · warning: {len(warnings)}")
        return 1 if errors else 0

    if args.command == "generate":
        targets = [
            (REGISTRY_JSON, json.dumps(data, ensure_ascii=False, indent=2) + "\n"),
            (PROJECT_GRAPH_JSON,
             json.dumps(build_graph(registry), ensure_ascii=False, indent=2) + "\n"),
        ]
        stale = []
        for path, content in targets:
            name = os.path.relpath(path, REPO)
            if write_if_needed(path, content, args.check):
                stale.append(name)
                if not args.check:
                    print(f"scritto {name}")
            else:
                print(f"{os.path.basename(path)} gia' allineato")
        if args.check and stale:
            for name in stale:
                print(f"non e' allineato alla sorgente: {name}")
            return 1
        return 0

    # `wiki` scrive **nel repository**, e dopo D-076 nel repository restano solo `docs/characters/` e
    # il blocco della roadmap: le pagine di gioco vivono nel clone e le aggiorna `deploy`. La pagina
    # `Stato-delle-feature` e' passata anch'essa al deploy, che la genera direttamente dove si legge.
    if args.command == "wiki":
        changed, stale = apply_wiki_blocks(data, args.check)
        roadmap_changed = apply_roadmap_block(data, args.check)
        for note in stale:
            print(f"WARN  {note}")
        if args.check and (changed or roadmap_changed):
            for ref in changed:
                print(f"disallineato: {ref}")
            if roadmap_changed:
                print(f"disallineato: {os.path.relpath(ROADMAP_V01, REPO)}")
            return 1
        for ref in changed:
            print(f"aggiornato {ref}")
        if roadmap_changed:
            print(f"aggiornato {os.path.relpath(ROADMAP_V01, REPO)}")
        if not (changed or roadmap_changed):
            print("pagine gia' allineate")
        return 0

    if args.command == "suite":
        count, files = suite_measure()

        # `--run-log`: il conteggio smette di essere un numero solo e diventa «N eseguiti su M
        # dichiarati». Non tocca i documenti generati — misura una RUN, che non e' uno stato del
        # repository — e non richiede `--check`: se un test dichiarato non e' stato eseguito, la
        # domanda non e' se scrivere un file, e' che la run non vale.
        if args.run_log:
            if not os.path.isfile(args.run_log):
                print(f"log non trovato: {args.run_log}")
                return 1
            scope, executed, missing, extra = suite_run_diff(args.run_log)
            print(f"run: {len(executed)} eseguiti su {len(scope)} dichiarati nel perimetro")
            for name in extra:
                print(f"WARN  eseguito ma non dichiarato in Source/RefactorTactics/Tests/: {name}")
            if not executed:
                print("nessun test eseguito: la run e' morta prima, oppure il filtro non prende nulla")
                return 1
            if missing:
                for name in missing:
                    print(f"MANCANTE dichiarato e non eseguito: {name}")
                print(f"{len(missing)} test dichiarati non sono stati eseguiti: la run non e' completa,")
                print("e zero falliti non vuol dire zero buchi (#486).")
                return 1
            print("nessun buco: ogni test dichiarato nel perimetro e' stato eseguito")
            return 0

        changed = apply_suite_count(args.check)
        print(f"suite: {count} test unici in {files} file su {current_head()}")
        if args.check and changed:
            for ref in changed:
                print(f"disallineato: {ref}")
            return 1
        for ref in changed:
            print(f"aggiornato {ref}")
        if not changed:
            print("conteggio gia' allineato")
        return 0

    if args.command == "shortlist":
        changed, missing = apply_shortlist(data, args.check)
        for note in missing:
            print(f"WARN  {note}")
        if args.check and changed:
            for ref in changed:
                print(f"disallineato: {ref}")
            return 1
        for ref in changed:
            print(f"aggiornato {ref}")
        if not changed:
            print("shortlist gia' allineate")
        return 0

    if args.command == "workbook":
        rows, error = apply_workbook(data, args.check, args.wiki_root)
        if error:
            print(error)
            return 1
        if not rows:
            print(f"{REFS_SHEET} gia' allineata")
            return 0
        if args.check:
            print(f"{REFS_SHEET} non e' allineata: {len(rows)} righe attese")
            return 1
        print(f"scritta {REFS_SHEET}: {len(rows)} relazioni")
        return 0

    if args.command == "deploy":
        if not args.wiki_root:
            print("deploy richiede --wiki-root <path del clone della Wiki>")
            return 1
        if not os.path.isdir(args.wiki_root):
            print(f"--wiki-root non e' una directory: {args.wiki_root}")
            return 1
        changed, missing = apply_wiki_deploy(data, args.wiki_root, check=not args.write)
        # Una sorgente che non raggiunge il clone e' un **errore**, non un avviso: il gate `ui_wiki`
        # della feature dice «spiegata all'utente», ma il lettore legge la Wiki pubblicata e li' la
        # pagina non c'e'. Il registry sta affermando una copertura falsa, non segnalando una lacuna —
        # stessa famiglia dello scenario orfano, che per questo e' bloccante (vedi feature-registry.md).
        for note in missing:
            print(f"ERROR {note}")
        verb = "aggiornata" if args.write else "da aggiornare"
        for name in changed:
            print(f"{verb}: {name}")
        print(f"\npagine {verb}: {len(changed)}")
        if not args.write and changed:
            print("sola lettura: aggiungi --write per scrivere nel clone della Wiki")
        if missing:
            print(f"sorgenti che non raggiungono il clone: {len(missing)} — il deploy e' incompleto")
            return 1
        if args.check and changed:
            return 1
        return 0

    if args.command == "report":
        print(render_audit(data))
        return 0

    return 0


if __name__ == "__main__":
    sys.exit(main())
