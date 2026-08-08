#!/usr/bin/env python3
"""Genera il workbook delle matrici stati/trasformazioni dal markdown che ne e' la fonte.

Il progetto ha tre `.xlsx` e **nessuno script li legge**: sono dump di consultazione, e uno di essi e'
gia' stato declassato a `RESEARCH` da D-023, che ha spostato l'autorita' dei numeri sui cataloghi
markdown. Una matrice di design in formato binario ricreerebbe quel problema: non diffabile, non
revisionabile in PR, invisibile ai gate.

Qui la fonte e' `docs/characters/matrici-stati-personaggio.md`; il foglio e' un **derivato**, non
versionato. E' il requisito §14B del prompt di consolidamento — `RecommendedCandidate`, `DesignStatus`
e `TargetVersion` non possono divergere fra documento e workbook perche' esiste una sola copia scritta
a mano.

Uso:
    python scripts/build-state-matrices-xlsx.py [destinazione.xlsx]

Default: `build/RefactorTactics_State_Matrices.xlsx` (cartella ignorata da git).
"""
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    sys.exit("Serve openpyxl:  pip install openpyxl")

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SORGENTE = os.path.join(REPO, "docs", "characters", "matrici-stati-personaggio.md")
DEFAULT_OUT = os.path.join(REPO, "build", "RefactorTactics_State_Matrices.xlsx")

# Un titolo di foglio Excel: max 31 caratteri, niente []:*?/\
VIETATI = re.compile(r"[\[\]:*?/\\]")


def _celle(riga):
    return [c.strip() for c in riga.strip().strip("|").split("|")]


def _e_separatore(riga):
    return bool(re.fullmatch(r"\|[\s:|-]+\|", riga.strip()))


def tabelle(testo):
    """Ogni tabella markdown, con il titolo della sezione (`##`/`###`) che la precede."""
    fuori = []
    sezione, sotto, righe = "Matrici", "", []
    in_fence = False

    for riga in testo.split("\n"):
        if riga.lstrip().startswith("```"):
            in_fence = not in_fence
        if in_fence:
            continue

        if riga.startswith("## "):
            sezione, sotto = riga[3:].strip(), ""
        elif riga.startswith("### "):
            sotto = riga[4:].strip()

        if riga.strip().startswith("|") and riga.strip().endswith("|"):
            if not _e_separatore(riga):
                righe.append(_celle(riga))
        elif righe:
            fuori.append((sezione, sotto, righe))
            righe = []
    if righe:
        fuori.append((sezione, sotto, righe))
    return fuori


def nome_foglio(sezione, sotto, usati):
    base = sezione.split("—")[0].strip() or sezione
    if sotto:
        base = f"{base} {sotto.split('—')[0].strip()}"
    base = VIETATI.sub("-", base)[:31].strip() or "Foglio"
    nome, n = base, 2
    while nome in usati:
        suffisso = f"_{n}"
        nome, n = base[: 31 - len(suffisso)] + suffisso, n + 1
    usati.add(nome)
    return nome


def main():
    if not os.path.exists(SORGENTE):
        sys.exit(f"Sorgente mancante: {SORGENTE}")
    out = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUT

    trovate = tabelle(open(SORGENTE, encoding="utf-8").read())
    if not trovate:
        sys.exit("Nessuna tabella trovata nel markdown: la fonte e' cambiata di forma?")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    intestazione = Font(bold=True, color="FFFFFF")
    sfondo = PatternFill("solid", fgColor="404040")
    usati = set()

    for sezione, sotto, righe in trovate:
        ws = wb.create_sheet(nome_foglio(sezione, sotto, usati))
        for r in righe:
            ws.append(r)
        for cella in ws[1]:
            cella.font, cella.fill = intestazione, sfondo
            cella.alignment = Alignment(vertical="center", wrap_text=True)
        for col in ws.columns:
            larghezza = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(larghezza + 2, 10), 60)
        ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"Scritto {out}")
    print(f"{len(trovate)} fogli, dalla fonte docs/characters/matrici-stati-personaggio.md")
    print("Il file e' un DERIVATO: non versionarlo, non modificarlo a mano.")


if __name__ == "__main__":
    sys.exit(main())
